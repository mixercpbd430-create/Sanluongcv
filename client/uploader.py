"""
Client Uploader — Gửi dữ liệu sản lượng, NVVH, LOSS lên server.
Chạy trên mỗi PC (Mixer, Pellet Feedmill, Pellet Mini).
Đọc file Excel local → parse dữ liệu → POST lên server.

Sử dụng:
    python uploader.py
"""

import os
import sys
import re
import json
import math
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import threading

# Try to import dependencies
try:
    import pandas as pd
except ImportError:
    print("Cần cài pandas: pip install pandas openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Cần cài requests: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Cần cài openpyxl: pip install openpyxl")
    sys.exit(1)


# ─── Config ────────────────────────────────────────────────
# When running as PyInstaller EXE, __file__ points to a temp dir.
# Use the directory of the actual EXE so config.json is persisted.
if getattr(sys, 'frozen', False):
    _app_dir = os.path.dirname(sys.executable)
else:
    _app_dir = os.path.dirname(os.path.abspath(__file__))

# Support --profile argument for per-user config files
_profile = ""
for i, arg in enumerate(sys.argv):
    if arg == "--profile" and i + 1 < len(sys.argv):
        _profile = sys.argv[i + 1]

if _profile:
    CONFIG_FILE = os.path.join(_app_dir, f"config_{_profile}.json")
else:
    CONFIG_FILE = os.path.join(_app_dir, "config.json")

DEFAULT_CONFIG = {
    "server_url": "https://api.binhduongfeedmill.com",
    "username": "",
    "password": "",
    "folder": "",
}

# File type configurations (same as data_loader.py)
FILE_CONFIGS = {
    "PL": {
        "pattern": "PL*.xlsx",
        "regex": r"(PL\d+)\s+(\d+)\.(\d+)\.xlsx",
        "sheet_name": "SAN LUONG (2)",
        "usecols": "A:E",
        "skiprows": 8,
        "category": "Pellet Mill",
    },
    "MIXER": {
        "pattern": "MIXER*.xls*",
        "regex": r"(MIXER)\s+T(\d+)\.(\d+)\.(xlsx|xlsm)",
        "sheet_name": "SAN LUONG",
        "usecols": "B:F",
        "skiprows": 7,
        "category": "Mixer",
        "extra_col": "AL",
    },
}

# All file types and lines — any user can upload everything
ALL_FILE_TYPES = ["PL", "MIXER"]
ALL_LINES = ["MIXER", "PL1", "PL2", "PL3", "PL4", "PL5", "PL6", "PL7"]


# ─── Config Load/Save ─────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            # Merge with defaults
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
    return DEFAULT_CONFIG.copy()


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ─── Helpers ───────────────────────────────────────────────
def safe_float(val, default=0.0):
    """Convert value to float, replacing NaN/inf with default."""
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


# ─── Excel Parser ──────────────────────────────────────────
def parse_excel_file(filepath, config):
    """Parse a single Excel file. Returns list of day entries."""
    filename = os.path.basename(filepath)
    match = re.match(config["regex"], filename, re.IGNORECASE)
    if not match:
        return None, None

    line_name = match.group(1).upper()
    month = int(match.group(2))
    year = int(match.group(3))

    try:
        df = pd.read_excel(
            filepath,
            sheet_name=config["sheet_name"],
            header=None,
            usecols=config["usecols"],
            skiprows=config["skiprows"],
            nrows=31,
        )
        df.columns = ["day", "ca1", "ca2", "ca3", "total"]
        df = df.fillna(0)

        entries = []
        for _, row in df.iterrows():
            day_num = int(row["day"]) if row["day"] else 0
            if day_num < 1 or day_num > 31:
                continue
            entry = {
                "line_name": line_name,
                "category": config["category"],
                "year": year,
                "month": month,
                "day": day_num,
                "ca1": round(safe_float(row["ca1"]), 2),
                "ca2": round(safe_float(row["ca2"]), 2),
                "ca3": round(safe_float(row["ca3"]), 2),
                "total": round(safe_float(row["total"]), 2),
                "cam_bot": 0,
            }
            entries.append(entry)

        # Read extra column (cam_bot) for MIXER
        extra_col = config.get("extra_col")
        if extra_col:
            try:
                df_extra = pd.read_excel(
                    filepath,
                    sheet_name=config["sheet_name"],
                    header=None,
                    usecols=extra_col,
                    skiprows=config["skiprows"],
                    nrows=31,
                )
                for idx, row in df_extra.iterrows():
                    if idx < len(entries):
                        val = row.iloc[0]
                        entries[idx]["cam_bot"] = round(safe_float(val), 2)
            except Exception:
                pass

        return f"{line_name} T{month}.{year}", entries

    except Exception as e:
        return filename, f"Lỗi: {e}"


# ─── NVVH Parser ───────────────────────────────────────────
def parse_nvvh_from_file(filepath, month, year):
    """Parse NVVH (operator names) from PL or MIXER file.
    PL files:    Row 49, columns AH(33), AL(37), AO(40) in day sheets 1-31.
    MIXER files: Row 85, columns M(12), P(15), S(18) in day sheets 1-31.
                 Format: "CA1:name1 name2", "CA2:name1 + name2", "CA 3:name"
    Returns list of {year, month, day, ca, pl_group, names}.
    """
    filename = os.path.basename(filepath).upper()
    # Determine pl_group
    if 'MIXER' in filename:
        pl_group = "mixer"
    elif any(f"PL{i}" in filename for i in range(1, 6)):
        pl_group = "pl1_5"
    elif any(f"PL{i}" in filename for i in range(6, 8)):
        pl_group = "pl6_7"
    else:
        return []

    # Only read from PL1 (for pl1_5 group), PL6 (for pl6_7 group), or MIXER
    if pl_group == "pl1_5" and "PL1 " not in filename and "PL1." not in filename:
        return []
    if pl_group == "pl6_7" and "PL6 " not in filename and "PL6." not in filename:
        return []

    # MIXER: Row 85, columns M(12), P(15), S(18) — 0-based indices
    # PL:    Row 49, columns AH(33), AL(37), AO(40) — 0-based indices
    if pl_group == "mixer":
        nvvh_row = 85
        ca_cols = {"ca1": 12, "ca2": 15, "ca3": 18}  # M=12, P=15, S=18
    else:
        nvvh_row = 49
        ca_cols = {"ca1": 33, "ca2": 37, "ca3": 40}  # AH=33, AL=37, AO=40

    entries = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for day in range(1, 32):
            sheet_name = str(day)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=nvvh_row, max_row=nvvh_row, values_only=False):
                for ca_key, col_idx in ca_cols.items():
                    if col_idx < len(row):
                        raw = row[col_idx].value
                        if raw:
                            names_str = str(raw).strip()
                            if pl_group == "mixer":
                                # Strip "CA1:", "CA2:", "CA 3:" prefix
                                names_str = re.sub(r'^CA\s*\d\s*[:：]\s*', '', names_str)
                                # Split by + or space-separated pairs
                                parts = re.split(r'[+]', names_str)
                                names_str = ','.join(n.strip() for n in parts if n.strip())
                            elif pl_group == "pl1_5":
                                # Split names by + or -
                                parts = re.split(r'[+\-]', names_str)
                                names_str = ','.join(n.strip() for n in parts if n.strip())
                            entries.append({
                                "year": year, "month": month,
                                "day": day, "ca": ca_key,
                                "pl_group": pl_group, "names": names_str,
                            })
        wb.close()
    except Exception as e:
        print(f"⚠️ NVVH parse error ({filepath}): {e}")

    return entries


# ─── LOSS Parser ───────────────────────────────────────────
def parse_loss_from_file(filepath, month, year):
    """Parse LOSS sheet from a PL file.
    Returns list of {year, month, day, pl_num, code, description, ca1_count, ...}.
    """
    filename = os.path.basename(filepath).upper()
    # Extract PL number or detect MIXER
    if 'MIXER' in filename:
        pl_num = 0  # Mixer uses pl_num=0
    else:
        pl_match = re.search(r'PL(\d+)', filename)
        if not pl_match:
            return []
        pl_num = int(pl_match.group(1))

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "LOSS" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["LOSS"]
        # Read all rows once into memory for performance
        all_rows = list(ws.iter_rows(min_row=7, max_row=450, values_only=False))
        wb.close()

        for day in range(1, 32):
            start_col = 5 + (day - 1) * 6  # 1-indexed
            for row in all_rows:
                code_cell = row[1] if len(row) > 1 else None
                desc_cell = row[2] if len(row) > 2 else None
                code = code_cell.value if code_cell else None
                if code is None:
                    continue

                desc = str(desc_cell.value) if desc_cell and desc_cell.value else ""

                vals = []
                for offset in range(6):
                    col_idx = start_col - 1 + offset
                    v = row[col_idx].value if col_idx < len(row) else None
                    try:
                        vals.append(float(v) if v else 0)
                    except (ValueError, TypeError):
                        vals.append(0)

                if any(v > 0 for v in vals):
                    entries.append({
                        "year": year, "month": month,
                        "day": day, "pl_num": pl_num,
                        "code": str(code), "description": desc,
                        "ca1_count": int(vals[0]),
                        "ca1_time": round(vals[1] * 60),
                        "ca2_count": int(vals[2]),
                        "ca2_time": round(vals[3] * 60),
                        "ca3_count": int(vals[4]),
                        "ca3_time": round(vals[5] * 60),
                    })
    except Exception as e:
        print(f"⚠️ LOSS parse error ({filepath}): {e}")

    return entries


def read_all_files(folder, username):
    """Read all Excel files from folder.
    Returns: (production_entries, nvvh_entries, loss_entries, file_results)
    """
    all_entries = []
    nvvh_entries = []
    loss_entries = []
    file_results = []

    for type_key in ALL_FILE_TYPES:
        config = FILE_CONFIGS[type_key]
        pattern = os.path.join(folder, config["pattern"])
        files = sorted(glob.glob(pattern))

        for filepath in files:
            label, result = parse_excel_file(filepath, config)
            if result is None:
                continue

            if isinstance(result, str):
                file_results.append(f"❌ {label}: {result}")
                continue

            # Accept all lines
            if result:
                all_entries.extend(result)
                total = sum(e["total"] for e in result)
                file_results.append(f"✅ {label}: {len(result)} ngày, {total:,.1f} tấn")

            # Parse NVVH + LOSS for PL and MIXER files
            if type_key == "PL" or type_key == "MIXER":
                fn = os.path.basename(filepath)
                match = re.match(config["regex"], fn, re.IGNORECASE)
                if match:
                    month = int(match.group(2))
                    year = int(match.group(3))
                    # NVVH (from PL1, PL6, or MIXER)
                    nvvh = parse_nvvh_from_file(filepath, month, year)
                    if nvvh:
                        nvvh_entries.extend(nvvh)
                        file_results.append(f"   👤 NVVH: {len(nvvh)} records")
                    # LOSS
                    loss = parse_loss_from_file(filepath, month, year)
                    if loss:
                        loss_entries.extend(loss)
                        file_results.append(f"   📋 LOSS: {len(loss)} records")

    return all_entries, nvvh_entries, loss_entries, file_results


def upload_data(server_url, username, password, entries, nvvh_entries=None, loss_entries=None):
    """Send production, NVVH, and LOSS data to server via API."""
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    url = f"{server_url.rstrip('/')}/api/upload-data"
    payload = {
        "username": username,
        "password": password,
        "entries": entries,
        "nvvh_entries": nvvh_entries or [],
        "loss_entries": loss_entries or [],
    }
    resp = requests.post(url, json=payload, timeout=300, verify=False)
    return resp.json()


# ─── GUI ───────────────────────────────────────────────────
class UploaderApp:
    def __init__(self):
        self.config = load_config()
        self.root = tk.Tk()
        title = "📊 Cập Nhật Sản Lượng"
        if _profile:
            title += f" — {_profile.upper()}"
        self.root.title(title)
        self.root.geometry("520x520")
        self.root.resizable(False, False)
        self.root.configure(bg="#1a1a2e")
        self._build_ui()

    def _build_ui(self):
        bg = "#1a1a2e"
        fg = "#e2e8f0"
        accent = "#6366f1"
        entry_bg = "#16213e"

        # Title
        tk.Label(self.root, text="📊 Cập Nhật Sản Lượng",
                 font=("Segoe UI", 16, "bold"), bg=bg, fg=fg).pack(pady=(15, 5))
        tk.Label(self.root, text="Đọc Excel → Gửi lên Server",
                 font=("Segoe UI", 9), bg=bg, fg="#94a3b8").pack()

        # Form frame
        form = tk.Frame(self.root, bg=bg)
        form.pack(pady=15, padx=25, fill="x")

        # Server URL
        tk.Label(form, text="Server URL:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=0, column=0, sticky="w", pady=4)
        self.server_var = tk.StringVar(value=self.config["server_url"])
        tk.Entry(form, textvariable=self.server_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=0, column=1, columnspan=2, pady=4, sticky="ew")

        # Username
        tk.Label(form, text="Username:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=1, column=0, sticky="w", pady=4)
        self.user_var = tk.StringVar(value=self.config["username"])
        tk.Entry(form, textvariable=self.user_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=1, column=1, columnspan=2, pady=4, sticky="ew")

        # Password
        tk.Label(form, text="Password:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=2, column=0, sticky="w", pady=4)
        self.pass_var = tk.StringVar(value=self.config.get("password", ""))
        tk.Entry(form, textvariable=self.pass_var, font=("Segoe UI", 10),
                 show="•", bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=2, column=1, columnspan=2, pady=4, sticky="ew")

        # Folder
        tk.Label(form, text="Folder Excel:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=3, column=0, sticky="w", pady=4)
        self.folder_var = tk.StringVar(value=self.config["folder"])
        tk.Entry(form, textvariable=self.folder_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=28).grid(row=3, column=1, pady=4, sticky="ew")
        tk.Button(form, text="📁", font=("Segoe UI", 10),
                  bg=accent, fg="white", relief="flat", width=3,
                  command=self._browse_folder).grid(row=3, column=2, padx=(5, 0))

        form.columnconfigure(1, weight=1)

        # Upload button
        self.btn_upload = tk.Button(
            self.root, text="🚀  Gửi dữ liệu lên hệ thống",
            font=("Segoe UI", 12, "bold"), bg=accent, fg="white",
            activebackground="#4f46e5", relief="flat", height=2,
            command=self._start_upload)
        self.btn_upload.pack(pady=10, padx=25, fill="x")

        # Status text
        self.status_text = tk.Text(
            self.root, height=10, bg="#0f172a", fg="#94a3b8",
            font=("Consolas", 9), relief="flat", state="disabled",
            wrap="word")
        self.status_text.pack(padx=25, pady=(0, 15), fill="both", expand=True)

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa file Excel")
        if folder:
            self.folder_var.set(folder)

    def _log(self, msg):
        self.status_text.config(state="normal")
        self.status_text.insert("end", msg + "\n")
        self.status_text.see("end")
        self.status_text.config(state="disabled")

    def _clear_log(self):
        self.status_text.config(state="normal")
        self.status_text.delete("1.0", "end")
        self.status_text.config(state="disabled")

    def _start_upload(self):
        """Start upload in background thread."""
        self.btn_upload.config(state="disabled", text="⏳ Đang xử lý...")
        self._clear_log()
        threading.Thread(target=self._do_upload, daemon=True).start()

    def _do_upload(self):
        try:
            server = self.server_var.get().strip()
            username = self.user_var.get().strip().lower()
            password = self.pass_var.get().strip()
            folder = self.folder_var.get().strip()

            if not all([server, username, password, folder]):
                self._log("❌ Vui lòng điền đầy đủ thông tin!")
                return

            if not os.path.isdir(folder):
                self._log(f"❌ Folder không tồn tại: {folder}")
                return

            # Save config (including password for auto mode)
            self.config["server_url"] = server
            self.config["username"] = username
            self.config["password"] = password
            self.config["folder"] = folder
            save_config(self.config)

            # Read Excel files
            self._log(f"📂 Đang đọc file Excel từ: {folder}")
            self._log(f"👤 User: {username}")
            self._log("")

            entries, nvvh_entries, loss_entries, file_results = read_all_files(folder, username)

            for msg in file_results:
                self._log(msg)

            if not entries and not nvvh_entries and not loss_entries:
                self._log("\n❌ Không tìm thấy dữ liệu phù hợp!")
                return

            self._log(f"\n📊 Tổng: {len(entries)} sản lượng, {len(nvvh_entries)} NVVH, {len(loss_entries)} LOSS")
            self._log(f"🚀 Đang gửi lên {server}...")

            # Upload
            result = upload_data(server, username, password, entries, nvvh_entries, loss_entries)

            if result.get("status") == "ok":
                now = datetime.now().strftime("%H:%M:%S")
                self._log(f"\n✅ Gửi thành công lúc {now}!")
                self._log(f"   Sản lượng: {result.get('inserted', 0)} bản ghi")
                if result.get('nvvh_count', 0) > 0:
                    self._log(f"   NVVH: {result['nvvh_count']} bản ghi")
                if result.get('loss_count', 0) > 0:
                    self._log(f"   LOSS: {result['loss_count']} bản ghi")
                if result.get("skipped", 0) > 0:
                    self._log(f"   Bỏ qua: {result['skipped']}")
            else:
                self._log(f"\n❌ Lỗi: {result.get('message', 'Unknown error')}")

            if result.get("errors"):
                for err in result["errors"]:
                    self._log(f"   ⚠️ {err}")

        except requests.exceptions.ConnectionError:
            self._log(f"\n❌ Không kết nối được đến server!")
            self._log(f"   Kiểm tra lại URL: {self.server_var.get()}")
        except Exception as e:
            self._log(f"\n❌ Lỗi: {str(e)}")
        finally:
            self.root.after(0, lambda: self.btn_upload.config(
                state="normal", text="🚀  Gửi dữ liệu lên hệ thống"))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = UploaderApp()
    app.run()
