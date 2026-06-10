"""
Client Uploader ALL — Gửi dữ liệu sản lượng MIXER + PL1~PL7, NVVH, LOSS lên server.
Upload tất cả dữ liệu lên https://api.binhduongfeedmill.com

Sử dụng:
    python uploader_all.py
    python uploader_all.py --profile <tên>
"""

import os
import sys
import re
import json
import math
import glob
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, date, timedelta
import threading

# Try to import dependencies
try:
    import pandas as pd
except ImportError:
    print("Cần cài pandas: pip install pandas openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Cần cài requests: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Cần cài openpyxl: pip install openpyxl")
    sys.exit(1)


# ─── Config ────────────────────────────────────────────────
# When running as PyInstaller EXE, __file__ points to a temp dir.
# Use the directory of the actual EXE so config.json is persisted.
if getattr(sys, 'frozen', False):
    _app_dir = os.path.dirname(sys.executable)
else:
    _app_dir = os.path.dirname(os.path.abspath(__file__))

# Support --profile argument for per-user config files
_profile = ""
for i, arg in enumerate(sys.argv):
    if arg == "--profile" and i + 1 < len(sys.argv):
        _profile = sys.argv[i + 1]

if _profile:
    CONFIG_FILE = os.path.join(_app_dir, f"config_all_{_profile}.json")
else:
    CONFIG_FILE = os.path.join(_app_dir, "config_all.json")

DEFAULT_CONFIG = {
    "server_url": "https://api.binhduongfeedmill.com",
    "local_url": "http://localhost:5009",
    "username": "",
    "password": "",
    "folder": "",
    "sp_mixer_url": "https://cpvn.sharepoint.com/sites/CPVDocument/BDG%20AgroFeed%20Production/Forms/AllItems.aspx?id=%2Fsites%2FCPVDocument%2FBDG%20AgroFeed%20Production%2FFile%20d%C3%B9ng%20chung%2FB%C3%81O%20C%C3%81O%20V%E1%BA%ACN%20H%C3%80NH%20S%E1%BA%A2N%20XU%E1%BA%A4T%2FB%C3%81O%20C%C3%81O%20V%E1%BA%ACN%20H%C3%80NH%20PHA%20TR%E1%BB%98N%2F2026",
    "sp_pellet_url": "https://cpvn.sharepoint.com/sites/CPVDocument/BDG%20AgroFeed%20Production/Forms/AllItems.aspx?id=%2Fsites%2FCPVDocument%2FBDG%20AgroFeed%20Production%2FFile%20d%C3%B9ng%20chung%2FB%C3%81O%20C%C3%81O%20V%E1%BA%ACN%20H%C3%80NH%20S%E1%BA%A2N%20XU%E1%BA%A4T%2FB%C3%81O%20C%C3%81O%20%20V%E1%BA%ACN%20H%C3%80NH%20C%C3%81M%20VI%C3%8AN%2F2026",
}

# ─── SharePoint Module ────────────────────────────────────
try:
    from sp_downloader import SPDownloader, HAS_PLAYWRIGHT
except ImportError:
    SPDownloader = None
    HAS_PLAYWRIGHT = False

# File type configurations — MIXER + PL (all lines)
FILE_CONFIGS = {
    "PL": {
        "pattern": "PL*.xlsx",
        "regex": r"(PL\d+)\s+(\d+)\.(\d+)\.xlsx",
        "sheet_name": "SAN LUONG (2)",
        "usecols": "A:E",
        "skiprows": 8,
        "category": "Pellet Mill",
    },
    "MIXER": {
        "pattern": "MIXER*.xls*",
        "regex": r"(MIXER)\s+T(\d+)\.(\d+)\.(xlsx|xlsm)",
        "sheet_name": "SAN LUONG",
        "usecols": "B:F",
        "skiprows": 7,
        "category": "Mixer",
        "extra_col": "AL",
    },
}

# All file types and lines
ALL_FILE_TYPES = ["PL", "MIXER"]
ALL_LINES = ["MIXER", "PL1", "PL2", "PL3", "PL4", "PL5", "PL6", "PL7"]


# ─── Config Load/Save ─────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            # Merge with defaults
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
    return DEFAULT_CONFIG.copy()


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ─── Helpers ───────────────────────────────────────────────
def safe_float(val, default=0.0):
    """Convert value to float, replacing NaN/inf with default."""
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


# ─── Excel Parser ──────────────────────────────────────────
def parse_excel_file(filepath, config):
    """Parse a single Excel file. Returns list of day entries."""
    filename = os.path.basename(filepath)
    match = re.match(config["regex"], filename, re.IGNORECASE)
    if not match:
        return None, None

    line_name = match.group(1).upper()
    month = int(match.group(2))
    year = int(match.group(3))

    try:
        df = pd.read_excel(
            filepath,
            sheet_name=config["sheet_name"],
            header=None,
            usecols=config["usecols"],
            skiprows=config["skiprows"],
            nrows=31,
        )
        df.columns = ["day", "ca1", "ca2", "ca3", "total"]
        df = df.fillna(0)

        entries = []
        for _, row in df.iterrows():
            day_num = int(row["day"]) if row["day"] else 0
            if day_num < 1 or day_num > 31:
                continue
            entry = {
                "line_name": line_name,
                "category": config["category"],
                "year": year,
                "month": month,
                "day": day_num,
                "ca1": round(safe_float(row["ca1"]), 2),
                "ca2": round(safe_float(row["ca2"]), 2),
                "ca3": round(safe_float(row["ca3"]), 2),
                "total": round(safe_float(row["total"]), 2),
                "cam_bot": 0,
            }
            entries.append(entry)

        # Read extra column (cam_bot) for MIXER
        extra_col = config.get("extra_col")
        if extra_col:
            try:
                df_extra = pd.read_excel(
                    filepath,
                    sheet_name=config["sheet_name"],
                    header=None,
                    usecols=extra_col,
                    skiprows=config["skiprows"],
                    nrows=31,
                )
                for idx, row in df_extra.iterrows():
                    if idx < len(entries):
                        val = row.iloc[0]
                        entries[idx]["cam_bot"] = round(safe_float(val), 2)
            except Exception:
                pass

        return f"{line_name} T{month}.{year}", entries

    except Exception as e:
        return filename, f"Lỗi: {e}"


# ─── NVVH Parser ───────────────────────────────────────────
def parse_nvvh_from_file(filepath, month, year):
    """Parse NVVH (operator names) from PL or MIXER file.
    PL files:    Row 49, columns AH(33), AL(37), AO(40) in day sheets 1-31.
    MIXER files: Row 85, columns M(12), P(15), S(18) in day sheets 1-31.
                 Format: "CA1:name1 name2", "CA2:name1 + name2", "CA 3:name"
    Returns list of {year, month, day, ca, pl_group, names}.
    """
    filename = os.path.basename(filepath).upper()
    # Determine pl_group
    if 'MIXER' in filename:
        pl_group = "mixer"
    elif any(f"PL{i}" in filename for i in range(1, 6)):
        pl_group = "pl1_5"
    elif any(f"PL{i}" in filename for i in range(6, 8)):
        pl_group = "pl6_7"
    else:
        return []

    # Only read from PL1 (for pl1_5 group), PL6 (for pl6_7 group), or MIXER
    if pl_group == "pl1_5" and "PL1 " not in filename and "PL1." not in filename:
        return []
    if pl_group == "pl6_7" and "PL6 " not in filename and "PL6." not in filename:
        return []

    # MIXER: Row 85, columns M(12), P(15), S(18) — 0-based indices
    # PL:    Row 49, columns AH(33), AL(37), AO(40) — 0-based indices
    if pl_group == "mixer":
        nvvh_row = 85
        ca_cols = {"ca1": 12, "ca2": 15, "ca3": 18}  # M=12, P=15, S=18
    else:
        nvvh_row = 49
        ca_cols = {"ca1": 33, "ca2": 37, "ca3": 40}  # AH=33, AL=37, AO=40

    entries = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for day in range(1, 32):
            sheet_name = str(day)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=nvvh_row, max_row=nvvh_row, values_only=False):
                for ca_key, col_idx in ca_cols.items():
                    if col_idx < len(row):
                        raw = row[col_idx].value
                        if raw:
                            names_str = str(raw).strip()
                            if pl_group == "mixer":
                                # Strip "CA1:", "CA2:", "CA 3:" prefix
                                names_str = re.sub(r'^CA\s*\d\s*[:：]\s*', '', names_str)
                                # Split by + or space-separated pairs
                                parts = re.split(r'[+]', names_str)
                                names_str = ','.join(n.strip() for n in parts if n.strip())
                            elif pl_group == "pl1_5":
                                # Split names by + or -
                                parts = re.split(r'[+\-]', names_str)
                                names_str = ','.join(n.strip() for n in parts if n.strip())
                            entries.append({
                                "year": year, "month": month,
                                "day": day, "ca": ca_key,
                                "pl_group": pl_group, "names": names_str,
                            })
        wb.close()
    except Exception as e:
        print(f"⚠️ NVVH parse error ({filepath}): {e}")

    return entries


# ─── LOSS Parser ───────────────────────────────────────────
def parse_loss_from_file(filepath, month, year):
    """Parse LOSS sheet from a PL or MIXER file.
    Returns list of {year, month, day, pl_num, code, description, ca1_count, ...}.
    """
    filename = os.path.basename(filepath).upper()
    # Extract PL number or detect MIXER
    if 'MIXER' in filename:
        pl_num = 0  # Mixer uses pl_num=0
    else:
        pl_match = re.search(r'PL(\d+)', filename)
        if not pl_match:
            return []
        pl_num = int(pl_match.group(1))

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "LOSS" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["LOSS"]
        # Read all rows once into memory for performance
        all_rows = list(ws.iter_rows(min_row=7, max_row=450, values_only=False))
        wb.close()

        for day in range(1, 32):
            start_col = 5 + (day - 1) * 6  # 1-indexed
            for row in all_rows:
                code_cell = row[1] if len(row) > 1 else None
                desc_cell = row[2] if len(row) > 2 else None
                code = code_cell.value if code_cell else None
                if code is None:
                    continue

                desc = str(desc_cell.value) if desc_cell and desc_cell.value else ""

                vals = []
                for offset in range(6):
                    col_idx = start_col - 1 + offset
                    v = row[col_idx].value if col_idx < len(row) else None
                    try:
                        vals.append(float(v) if v else 0)
                    except (ValueError, TypeError):
                        vals.append(0)

                if any(v > 0 for v in vals):
                    # MIXER time is already in minutes; PL time is in hours → convert to minutes
                    time_mult = 1 if pl_num == 0 else 60
                    entries.append({
                        "year": year, "month": month,
                        "day": day, "pl_num": pl_num,
                        "code": str(code), "description": desc,
                        "ca1_count": int(vals[0]),
                        "ca1_time": round(vals[1] * time_mult),
                        "ca2_count": int(vals[2]),
                        "ca2_time": round(vals[3] * time_mult),
                        "ca3_count": int(vals[4]),
                        "ca3_time": round(vals[5] * time_mult),
                    })
    except Exception as e:
        print(f"⚠️ LOSS parse error ({filepath}): {e}")

    return entries


# ─── Khuôn Parser ─────────────────────────────────────────
def parse_khuon_from_file(filepath, pl_num, month, year):
    """Parse SERI sheet from a PL file for khuon tracking data.
    Returns list of {
        pl_num, year, month, seri, thong_so,
        days: {1: val, 2: val, ...},
        tong_thang, ton_truoc, tong
    }
    """
    # Column layout differs by PL:
    #   PL2: seri=C(2), khuon=D(3), day1=E(4)
    #   Others: seri=B(1), khuon=C(2), day1=D(3)
    if pl_num == 2:
        seri_idx, khuon_idx, day1_idx = 2, 3, 4
    else:
        seri_idx, khuon_idx, day1_idx = 1, 2, 3

    # Summary columns: day1 + 31, day1 + 32, day1 + 33
    tong_thang_idx = day1_idx + 31
    ton_truoc_idx = day1_idx + 32
    tong_idx = day1_idx + 33

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if 'SERI' not in wb.sheetnames:
            wb.close()
            return []

        ws = wb['SERI']
        for row in ws.iter_rows(min_row=3, max_row=100, min_col=1,
                                max_col=tong_idx + 1, values_only=True):
            seri_val = row[seri_idx] if len(row) > seri_idx else None
            if seri_val is None or str(seri_val).strip() == '':
                continue

            thong_so = row[khuon_idx] if len(row) > khuon_idx else ''

            # Read daily values (day 1-31) — only non-zero to reduce payload
            days = {}
            for d in range(31):
                col_idx = day1_idx + d
                val = row[col_idx] if len(row) > col_idx else 0
                v = round(float(val), 2) if val else 0
                if v > 0:
                    days[d + 1] = v

            # Summary columns
            tong_thang = row[tong_thang_idx] if len(row) > tong_thang_idx else 0
            ton_truoc = row[ton_truoc_idx] if len(row) > ton_truoc_idx else 0
            tong = row[tong_idx] if len(row) > tong_idx else 0

            entries.append({
                'pl_num': pl_num,
                'year': year,
                'month': month,
                'seri': str(seri_val).strip(),
                'thong_so': str(thong_so).strip() if thong_so else '',
                'days': days,
                'tong_thang': round(float(tong_thang), 2) if tong_thang else 0,
                'ton_truoc': round(float(ton_truoc), 2) if ton_truoc else 0,
                'tong': round(float(tong), 2) if tong else 0,
            })

        wb.close()
    except Exception as e:
        print(f"⚠️ Khuôn parse error ({filepath}): {e}")

    return entries


def read_all_files(folder, username):
    """Read all Excel files from folder (MIXER + PL1~PL7).
    Returns: (production_entries, nvvh_entries, loss_entries, khuon_entries, file_results)
    """
    all_entries = []
    nvvh_entries = []
    loss_entries = []
    khuon_entries = []
    file_results = []

    for type_key in ALL_FILE_TYPES:
        config = FILE_CONFIGS[type_key]
        pattern = os.path.join(folder, config["pattern"])
        files = sorted(glob.glob(pattern))

        for filepath in files:
            label, result = parse_excel_file(filepath, config)
            if result is None:
                continue

            if isinstance(result, str):
                file_results.append(f"❌ {label}: {result}")
                continue

            # Accept all lines
            if result:
                all_entries.extend(result)
                total = sum(e["total"] for e in result)
                file_results.append(f"✅ {label}: {len(result)} ngày, {total:,.1f} tấn")

            # Parse NVVH + LOSS for PL and MIXER files
            if type_key == "PL" or type_key == "MIXER":
                fn = os.path.basename(filepath)
                match = re.match(config["regex"], fn, re.IGNORECASE)
                if match:
                    month = int(match.group(2))
                    year = int(match.group(3))
                    # NVVH (from PL1, PL6, or MIXER)
                    nvvh = parse_nvvh_from_file(filepath, month, year)
                    if nvvh:
                        nvvh_entries.extend(nvvh)
                        file_results.append(f"   👤 NVVH: {len(nvvh)} records")
                    # LOSS
                    loss = parse_loss_from_file(filepath, month, year)
                    if loss:
                        loss_entries.extend(loss)
                        file_results.append(f"   📋 LOSS: {len(loss)} records")

            # Parse KHUÔN from PL files only (SERI sheet)
            if type_key == "PL":
                fn = os.path.basename(filepath)
                match = re.match(config["regex"], fn, re.IGNORECASE)
                if match:
                    line_name = match.group(1).upper()
                    pl_num = int(re.search(r'\d+', line_name).group())
                    month = int(match.group(2))
                    year = int(match.group(3))
                    khuon = parse_khuon_from_file(filepath, pl_num, month, year)
                    if khuon:
                        khuon_entries.extend(khuon)
                        file_results.append(f"   🔩 KHUÔN: {len(khuon)} khuôn")

    return all_entries, nvvh_entries, loss_entries, khuon_entries, file_results


def _post_with_retry(url, payload, max_retries=3, timeout=90):
    """POST JSON with retry and exponential backoff.
    Returns (response_json_dict, error_string_or_None).
    """
    import urllib3
    import time as _time
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    last_err = None
    for attempt in range(max_retries):
        try:
            resp = requests.post(url, json=payload, timeout=timeout, verify=False)
            try:
                return resp.json(), None
            except (json.JSONDecodeError, ValueError):
                body_preview = resp.text[:200].strip() if resp.text else "(empty)"
                last_err = f"HTTP {resp.status_code} — không phải JSON. Nội dung: {body_preview}"
                # Retry on 5xx / Cloudflare errors
                if resp.status_code >= 500:
                    _time.sleep(3 * (attempt + 1))
                    continue
                return {"status": "error", "message": last_err}, last_err
        except requests.exceptions.Timeout:
            last_err = f"Timeout sau {timeout}s (lần {attempt + 1}/{max_retries})"
            _time.sleep(3 * (attempt + 1))
        except requests.exceptions.ConnectionError as ce:
            last_err = f"Không kết nối được: {ce}"
            _time.sleep(3 * (attempt + 1))
        except Exception as e:
            last_err = str(e)
            break

    return {"status": "error", "message": last_err}, last_err


def _group_entries_by_line_month(entries):
    """Group production entries by (line_name, year, month) for batching."""
    groups = {}
    for e in entries:
        key = (e.get("line_name", ""), e.get("year", 0), e.get("month", 0))
        groups.setdefault(key, []).append(e)
    return groups


def _group_loss_by_pl_month(loss_entries):
    """Group LOSS entries by (pl_num, year, month) for batching."""
    groups = {}
    for lo in loss_entries:
        key = (lo.get("pl_num", 0), lo.get("year", 0), lo.get("month", 0))
        groups.setdefault(key, []).append(lo)
    return groups


def _group_nvvh_by_group_month(nvvh_entries):
    """Group NVVH entries by (pl_group, year, month) for batching."""
    groups = {}
    for nv in nvvh_entries:
        key = (nv.get("pl_group", ""), nv.get("year", 0), nv.get("month", 0))
        groups.setdefault(key, []).append(nv)
    return groups


def _filter_recent_days(entries, nvvh_entries, loss_entries, recent_days=3):
    """Filter entries/nvvh/loss to only include the N most recent days.
    Based on today's date, keeps data for (today - recent_days) to (today - 1).
    For example, if today is 10/6/2026 and recent_days=3, keeps days 7, 8, 9 of June 2026.
    Also handles month boundaries (e.g., today is 2/7 → keeps 29, 30/6 + 1/7).
    """
    today = date.today()
    # Build set of (year, month, day) tuples for the recent window
    recent_dates = set()
    for i in range(1, recent_days + 1):
        d = today - timedelta(days=i)
        recent_dates.add((d.year, d.month, d.day))

    filtered_entries = [
        e for e in entries
        if (int(e.get("year", 0)), int(e.get("month", 0)), int(e.get("day", 0))) in recent_dates
    ]
    filtered_nvvh = [
        nv for nv in (nvvh_entries or [])
        if (int(nv.get("year", 0)), int(nv.get("month", 0)), int(nv.get("day", 0))) in recent_dates
    ]
    filtered_loss = [
        lo for lo in (loss_entries or [])
        if (int(lo.get("year", 0)), int(lo.get("month", 0)), int(lo.get("day", 0))) in recent_dates
    ]
    return filtered_entries, filtered_nvvh, filtered_loss, recent_dates


def upload_data(server_url, username, password, entries, nvvh_entries=None, loss_entries=None):
    """Send production, NVVH, and LOSS data to server via API (single request)."""
    url = f"{server_url.rstrip('/')}/api/upload-data"
    payload = {
        "username": username,
        "password": password,
        "entries": entries,
        "nvvh_entries": nvvh_entries or [],
        "loss_entries": loss_entries or [],
    }
    result, err = _post_with_retry(url, payload, max_retries=3, timeout=120)
    return result


def upload_data_batched(server_url, username, password, entries, nvvh_entries=None, loss_entries=None, log_fn=None):
    """Send data in batches grouped by line+month to avoid Cloudflare timeout.
    Returns aggregated result dict.
    """
    url = f"{server_url.rstrip('/')}/api/upload-data"
    nvvh_entries = nvvh_entries or []
    loss_entries = loss_entries or []

    # Build batches: each batch = 1 line+month of production + matching NVVH/LOSS
    prod_groups = _group_entries_by_line_month(entries)
    loss_groups = _group_loss_by_pl_month(loss_entries)
    nvvh_groups = _group_nvvh_by_group_month(nvvh_entries)

    # Map line_name → pl_num for matching LOSS groups
    def _line_to_pl(line_name):
        if line_name == "MIXER":
            return 0
        try:
            return int(re.search(r'\d+', line_name).group())
        except Exception:
            return -1

    # Map line_name → pl_group for matching NVVH groups
    def _line_to_nvvh_group(line_name):
        if line_name == "MIXER":
            return "mixer"
        pl = _line_to_pl(line_name)
        if 1 <= pl <= 5:
            return "pl1_5"
        if 6 <= pl <= 7:
            return "pl6_7"
        return None

    total_inserted = 0
    total_nvvh = 0
    total_loss = 0
    total_skipped = 0
    all_errors = []
    sent_nvvh_keys = set()
    sent_loss_keys = set()

    batch_num = 0
    total_batches = len(prod_groups)
    # If there are NVVH/LOSS groups not covered by production, add extra batches
    extra_loss_keys = set(loss_groups.keys())
    extra_nvvh_keys = set(nvvh_groups.keys())

    for (line_name, year, month), prod_batch in prod_groups.items():
        batch_num += 1
        pl_num = _line_to_pl(line_name)
        nvvh_group_name = _line_to_nvvh_group(line_name)

        # Attach matching LOSS
        loss_key = (pl_num, year, month)
        batch_loss = loss_groups.get(loss_key, [])
        if loss_key in extra_loss_keys:
            extra_loss_keys.discard(loss_key)
        sent_loss_keys.add(loss_key)

        # Attach matching NVVH (only once per group)
        batch_nvvh = []
        if nvvh_group_name:
            nvvh_key = (nvvh_group_name, year, month)
            if nvvh_key not in sent_nvvh_keys:
                batch_nvvh = nvvh_groups.get(nvvh_key, [])
                sent_nvvh_keys.add(nvvh_key)
                if nvvh_key in extra_nvvh_keys:
                    extra_nvvh_keys.discard(nvvh_key)

        if log_fn:
            log_fn(f"   📦 Batch {batch_num}/{total_batches}: {line_name} T{month}.{year} "
                   f"({len(prod_batch)} SL, {len(batch_loss)} LOSS, {len(batch_nvvh)} NVVH)")

        payload = {
            "username": username,
            "password": password,
            "entries": prod_batch,
            "nvvh_entries": batch_nvvh,
            "loss_entries": batch_loss,
        }
        result, err = _post_with_retry(url, payload, max_retries=3, timeout=90)

        if result.get("status") == "ok":
            total_inserted += result.get("inserted", 0)
            total_nvvh += result.get("nvvh_count", 0)
            total_loss += result.get("loss_count", 0)
            total_skipped += result.get("skipped", 0)
        else:
            msg = result.get("message", err or "Unknown")
            all_errors.append(f"{line_name} T{month}.{year}: {msg}")
            if log_fn:
                log_fn(f"   ❌ {line_name} T{month}.{year}: {msg}")

    # Send any remaining NVVH/LOSS not yet sent
    for nvvh_key in extra_nvvh_keys:
        if nvvh_key not in sent_nvvh_keys:
            batch_nvvh = nvvh_groups[nvvh_key]
            payload = {
                "username": username, "password": password,
                "entries": [], "nvvh_entries": batch_nvvh, "loss_entries": [],
            }
            result, _ = _post_with_retry(url, payload, max_retries=3, timeout=90)
            if result.get("status") == "ok":
                total_nvvh += result.get("nvvh_count", 0)

    for loss_key in extra_loss_keys:
        if loss_key not in sent_loss_keys:
            batch_loss = loss_groups[loss_key]
            payload = {
                "username": username, "password": password,
                "entries": [], "nvvh_entries": [], "loss_entries": batch_loss,
            }
            result, _ = _post_with_retry(url, payload, max_retries=3, timeout=90)
            if result.get("status") == "ok":
                total_loss += result.get("loss_count", 0)

    status = "ok" if not all_errors else "partial"
    return {
        "status": status,
        "inserted": total_inserted,
        "nvvh_count": total_nvvh,
        "loss_count": total_loss,
        "skipped": total_skipped,
        "errors": all_errors[:10],
    }


def upload_khuon(server_url, username, password, khuon_entries):
    """Send khuon tracking data to server via API (single request)."""
    url = f"{server_url.rstrip('/')}/api/upload-khuon"
    payload = {
        "username": username,
        "password": password,
        "khuon_entries": khuon_entries,
    }
    result, err = _post_with_retry(url, payload, max_retries=3, timeout=120)
    return result


def upload_khuon_batched(server_url, username, password, khuon_entries, log_fn=None):
    """Send khuon data in batches grouped by PL to avoid Cloudflare timeout."""
    url = f"{server_url.rstrip('/')}/api/upload-khuon"

    # Group by (pl_num, year, month)
    groups = {}
    for k in khuon_entries:
        key = (k.get("pl_num", 0), k.get("year", 0), k.get("month", 0))
        groups.setdefault(key, []).append(k)

    total_count = 0
    all_errors = []
    batch_num = 0
    total_batches = len(groups)

    for (pl_num, year, month), batch in groups.items():
        batch_num += 1
        if log_fn:
            log_fn(f"   📦 Khuôn batch {batch_num}/{total_batches}: PL{pl_num} T{month}.{year} ({len(batch)} khuôn)")

        payload = {
            "username": username,
            "password": password,
            "khuon_entries": batch,
        }
        result, err = _post_with_retry(url, payload, max_retries=3, timeout=90)

        if result.get("status") == "ok":
            total_count += result.get("khuon_count", 0)
        else:
            msg = result.get("message", err or "Unknown")
            all_errors.append(f"PL{pl_num} T{month}.{year}: {msg}")
            if log_fn:
                log_fn(f"   ❌ PL{pl_num} T{month}.{year}: {msg}")

    return {
        "status": "ok" if not all_errors else "partial",
        "khuon_count": total_count,
        "errors": all_errors[:10],
    }


# ─── GUI ───────────────────────────────────────────────────
class UploaderApp:
    def __init__(self):
        self.config = load_config()
        self.root = tk.Tk()
        title = "📊 Cập Nhật Sản Lượng — MIXER + PL1~PL7"
        if _profile:
            title += f" — {_profile.upper()}"
        self.root.title(title)
        self.root.geometry("560x760")
        self.root.resizable(False, False)
        self.root.configure(bg="#1a1a2e")
        self._sp_downloader = None
        self._build_ui()
        self._update_sp_status()

    def _get_sp_downloader(self):
        """Lazy init SPDownloader."""
        if self._sp_downloader is None and SPDownloader:
            self._sp_downloader = SPDownloader(
                log_fn=self._log,
                stop_check=lambda: False
            )
        return self._sp_downloader

    def _build_ui(self):
        bg = "#1a1a2e"
        fg = "#e2e8f0"
        accent = "#10b981"
        sp_accent = "#3b82f6"  # blue for SharePoint
        entry_bg = "#16213e"

        # Title
        tk.Label(self.root, text="📊 Cập Nhật Sản Lượng",
                 font=("Segoe UI", 16, "bold"), bg=bg, fg=fg).pack(pady=(15, 2))
        tk.Label(self.root, text="MIXER + PL1 ~ PL7",
                 font=("Segoe UI", 13, "bold"), bg=bg, fg=accent).pack(pady=(0, 2))
        tk.Label(self.root, text="Đọc Excel → Gửi lên api.binhduongfeedmill.com + Local",
                 font=("Segoe UI", 9), bg=bg, fg="#94a3b8").pack()

        # Form frame
        form = tk.Frame(self.root, bg=bg)
        form.pack(pady=15, padx=25, fill="x")

        # Server URL
        tk.Label(form, text="Server URL:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=0, column=0, sticky="w", pady=4)
        self.server_var = tk.StringVar(value=self.config["server_url"])
        tk.Entry(form, textvariable=self.server_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=38).grid(row=0, column=1, columnspan=2, pady=4, sticky="ew")

        # Local URL
        tk.Label(form, text="Local URL:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=1, column=0, sticky="w", pady=4)
        self.local_var = tk.StringVar(value=self.config.get("local_url", "http://localhost:5009"))
        tk.Entry(form, textvariable=self.local_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=38).grid(row=1, column=1, columnspan=2, pady=4, sticky="ew")

        # Username
        tk.Label(form, text="Username:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=2, column=0, sticky="w", pady=4)
        self.user_var = tk.StringVar(value=self.config["username"])
        tk.Entry(form, textvariable=self.user_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=38).grid(row=2, column=1, columnspan=2, pady=4, sticky="ew")

        # Password
        tk.Label(form, text="Password:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=3, column=0, sticky="w", pady=4)
        self.pass_var = tk.StringVar(value=self.config.get("password", ""))
        tk.Entry(form, textvariable=self.pass_var, font=("Segoe UI", 10),
                 show="•", bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=38).grid(row=3, column=1, columnspan=2, pady=4, sticky="ew")

        # Folder
        tk.Label(form, text="Folder Excel:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=4, column=0, sticky="w", pady=4)
        self.folder_var = tk.StringVar(value=self.config["folder"])
        tk.Entry(form, textvariable=self.folder_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=31).grid(row=4, column=1, pady=4, sticky="ew")
        tk.Button(form, text="📁", font=("Segoe UI", 10),
                  bg=accent, fg="white", relief="flat", width=3,
                  command=self._browse_folder).grid(row=4, column=2, padx=(5, 0))

        form.columnconfigure(1, weight=1)

        # Lines info
        lines_frame = tk.Frame(self.root, bg=bg)
        lines_frame.pack(padx=25, fill="x")
        tk.Label(lines_frame, text="Dây chuyền: ",
                 font=("Segoe UI", 9), bg=bg, fg="#94a3b8").pack(side="left")
        tk.Label(lines_frame, text="MIXER • PL1 • PL2 • PL3 • PL4 • PL5 • PL6 • PL7",
                 font=("Segoe UI", 9, "bold"), bg=bg, fg="#10b981").pack(side="left")

        # Upload button (folder local)
        self.btn_upload = tk.Button(
            self.root, text="🚀  Gửi tất cả dữ liệu lên hệ thống",
            font=("Segoe UI", 12, "bold"), bg=accent, fg="white",
            activebackground="#059669", relief="flat", height=2,
            command=self._start_upload)
        self.btn_upload.pack(pady=(10, 5), padx=25, fill="x")

        # ── SharePoint Section ────────────────────────────
        sp_frame = tk.Frame(self.root, bg="#111827", relief="flat", bd=0)
        sp_frame.pack(padx=25, fill="x", pady=(5, 0))

        # SP header + status
        sp_header = tk.Frame(sp_frame, bg="#111827")
        sp_header.pack(fill="x", padx=10, pady=(8, 4))
        tk.Label(sp_header, text="☁️ SharePoint",
                 font=("Segoe UI", 11, "bold"), bg="#111827", fg="#93c5fd").pack(side="left")
        self.sp_status_label = tk.Label(
            sp_header, text="⏳ Đang kiểm tra...",
            font=("Segoe UI", 9), bg="#111827", fg="#94a3b8")
        self.sp_status_label.pack(side="right")

        # SP buttons row
        sp_btns = tk.Frame(sp_frame, bg="#111827")
        sp_btns.pack(fill="x", padx=10, pady=(0, 8))

        self.btn_sp_login = tk.Button(
            sp_btns, text="🔑 Đăng nhập SP",
            font=("Segoe UI", 10, "bold"), bg=sp_accent, fg="white",
            activebackground="#2563eb", relief="flat",
            command=self._start_sp_login)
        self.btn_sp_login.pack(side="left", padx=(0, 5), expand=True, fill="x")

        self.btn_sp_download = tk.Button(
            sp_btns, text="📥 Tải SP → Gửi Local",
            font=("Segoe UI", 10, "bold"), bg="#8b5cf6", fg="white",
            activebackground="#7c3aed", relief="flat",
            command=self._start_sp_download)
        self.btn_sp_download.pack(side="left", padx=(5, 0), expand=True, fill="x")

        # Status text
        self.status_text = tk.Text(
            self.root, height=10, bg="#0f172a", fg="#94a3b8",
            font=("Consolas", 9), relief="flat", state="disabled",
            wrap="word")
        self.status_text.pack(padx=25, pady=(10, 15), fill="both", expand=True)

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa file Excel (MIXER + PL)")
        if folder:
            self.folder_var.set(folder)

    def _log(self, msg):
        self.status_text.config(state="normal")
        self.status_text.insert("end", msg + "\n")
        self.status_text.see("end")
        self.status_text.config(state="disabled")

    def _clear_log(self):
        self.status_text.config(state="normal")
        self.status_text.delete("1.0", "end")
        self.status_text.config(state="disabled")

    # ─── SharePoint methods ───────────────────────────────

    def _update_sp_status(self):
        """Cập nhật trạng thái đăng nhập SharePoint trên GUI."""
        dl = self._get_sp_downloader()
        if not SPDownloader:
            self.sp_status_label.config(text="⚠️ Chưa cài Playwright", fg="#f59e0b")
            return
        if dl and dl.check_login():
            self.sp_status_label.config(text="✅ Đã đăng nhập", fg="#10b981")
        else:
            self.sp_status_label.config(text="❌ Chưa đăng nhập", fg="#ef4444")

    def _start_sp_login(self):
        """Bắt đầu đăng nhập SharePoint trong thread riêng."""
        if not SPDownloader:
            self._log("❌ Chưa cài Playwright! Chạy: pip install playwright && playwright install chromium")
            return
        self.btn_sp_login.config(state="disabled", text="⏳ Đang mở browser...")
        self._clear_log()
        threading.Thread(target=self._do_sp_login, daemon=True).start()

    def _do_sp_login(self):
        """Thực hiện đăng nhập SharePoint."""
        try:
            dl = self._get_sp_downloader()
            sp_url = self.config.get("sp_mixer_url", "")
            if not sp_url:
                self._log("❌ Chưa có URL SharePoint trong config!")
                return
            dl.login_sharepoint(sp_url)
            self.root.after(0, self._update_sp_status)
        except Exception as e:
            self._log(f"❌ Lỗi đăng nhập SharePoint: {str(e)[:150]}")
        finally:
            self.root.after(0, lambda: self.btn_sp_login.config(
                state="normal", text="🔑 Đăng nhập SP"))

    def _start_sp_download(self):
        """Bắt đầu tải từ SharePoint + gửi Local trong thread riêng."""
        if not SPDownloader:
            self._log("❌ Chưa cài Playwright! Chạy: pip install playwright && playwright install chromium")
            return
        dl = self._get_sp_downloader()
        if not dl.check_login():
            self._log("❌ Chưa đăng nhập SharePoint! Nhấn '🔑 Đăng nhập SP' trước.")
            return
        self.btn_sp_download.config(state="disabled", text="⏳ Đang tải...")
        self.btn_sp_login.config(state="disabled")
        self.btn_upload.config(state="disabled")
        self._clear_log()
        threading.Thread(target=self._do_sp_download, daemon=True).start()

    def _do_sp_download(self):
        """Tải file từ SharePoint → parse → gửi localhost:5009."""
        try:
            local = self.local_var.get().strip()
            username = self.user_var.get().strip().lower()
            password = self.pass_var.get().strip()

            if not all([local, username, password]):
                self._log("❌ Vui lòng điền Username, Password và Local URL!")
                return

            # Tự lấy tháng/năm hiện tại
            now = datetime.now()
            month = now.month
            year = now.year

            self._log(f"☁️ Tải file từ SharePoint — Tháng {month}/{year}")
            self._log(f"👤 User: {username}")
            self._log(f"🏠 Gửi lên: {local}")
            self._log("")

            # Tải file từ SharePoint
            dl = self._get_sp_downloader()
            sp_mixer_url = self.config.get("sp_mixer_url", "")
            sp_pellet_url = self.config.get("sp_pellet_url", "")

            download_dir = dl.download_all_production(sp_mixer_url, sp_pellet_url, month, year)
            if not download_dir:
                self._log("\n❌ Không tải được file nào từ SharePoint!")
                return

            # Parse Excel từ thư mục đã tải
            self._log(f"\n📂 Đang đọc file Excel từ: {download_dir}")
            entries, nvvh_entries, loss_entries, khuon_entries, file_results = read_all_files(download_dir, username)

            for msg in file_results:
                self._log(msg)

            if not entries and not nvvh_entries and not loss_entries and not khuon_entries:
                self._log("\n❌ Không tìm thấy dữ liệu phù hợp trong file đã tải!")
                return

            self._log(f"\n📊 Tổng: {len(entries)} sản lượng, {len(nvvh_entries)} NVVH, "
                      f"{len(loss_entries)} LOSS, {len(khuon_entries)} KHUÔN")

            # Khởi động local server + gửi dữ liệu
            self._start_local_server()
            ok = self._upload_to_server("LOCAL", local, username, password,
                                         entries, nvvh_entries, loss_entries, khuon_entries)

            self._log(f"\n{'═' * 50}")
            if ok:
                self._log("🎉 Hoàn tất — Tải SharePoint + Gửi Local thành công!")
            else:
                self._log("⚠️ Hoàn tất — có lỗi khi gửi, xem chi tiết ở trên.")

        except Exception as e:
            self._log(f"\n❌ Lỗi: {str(e)}")
        finally:
            self.root.after(0, lambda: (
                self.btn_sp_download.config(state="normal", text="📥 Tải SP → Gửi Local"),
                self.btn_sp_login.config(state="normal", text="🔑 Đăng nhập SP"),
                self.btn_upload.config(state="normal", text="🚀  Gửi tất cả dữ liệu lên hệ thống"),
            ))

    def _start_local_server(self):
        """Launch start.bat in background to ensure local Flask server is running."""
        start_bat = r"D:\Github\SanluongcvReact1\start.bat"
        if not os.path.exists(start_bat):
            self._log(f"⚠️ Không tìm thấy {start_bat}")
            return
        try:
            self._log(f"🖥️ Đang khởi động local server (start.bat)...")
            # Launch in a new console window, non-blocking
            subprocess.Popen(
                ["cmd", "/c", "start", "Local Server", start_bat],
                cwd=r"D:\Github\SanluongcvReact1",
                shell=True,
            )
            # Wait a few seconds for server to start
            import time as _time
            self._log(f"⏳ Chờ server local khởi động (5 giây)...")
            _time.sleep(5)
            self._log(f"✅ Local server đã khởi động!")
        except Exception as e:
            self._log(f"⚠️ Lỗi khởi động start.bat: {e}")

    def _start_upload(self):
        """Start upload in background thread."""
        self.btn_upload.config(state="disabled", text="⏳ Đang xử lý...")
        self._clear_log()
        threading.Thread(target=self._do_upload, daemon=True).start()

    def _upload_to_server(self, label, server, username, password, entries, nvvh_entries, loss_entries, khuon_entries):
        """Upload all data to a single server. Returns True if successful.
        Uses batched upload for REMOTE (Cloudflare), single request for LOCAL.
        """
        self._log(f"\n{'═'*50}")
        self._log(f"🚀 Đang gửi lên {label}: {server}")
        self._log(f"{'═'*50}")
        success = True
        is_remote = label == "REMOTE"

        try:
            # Upload production + NVVH + LOSS
            if entries or nvvh_entries or loss_entries:
                if is_remote:
                    self._log(f"📡 [{label}] Gửi theo batch (tránh timeout Cloudflare)...")
                    result = upload_data_batched(
                        server, username, password, entries, nvvh_entries, loss_entries,
                        log_fn=self._log)
                else:
                    result = upload_data(server, username, password, entries, nvvh_entries, loss_entries)

                if result.get("status") in ("ok", "partial"):
                    now = datetime.now().strftime("%H:%M:%S")
                    self._log(f"✅ [{label}] Gửi sản lượng thành công lúc {now}!")
                    self._log(f"   Sản lượng: {result.get('inserted', 0)} bản ghi")
                    if result.get('nvvh_count', 0) > 0:
                        self._log(f"   NVVH: {result['nvvh_count']} bản ghi")
                    if result.get('loss_count', 0) > 0:
                        self._log(f"   LOSS: {result['loss_count']} bản ghi")
                    if result.get("skipped", 0) > 0:
                        self._log(f"   Bỏ qua: {result['skipped']}")
                    if result.get("status") == "partial":
                        self._log(f"   ⚠️ [{label}] Một số batch bị lỗi (xem chi tiết bên dưới)")
                        success = False
                else:
                    self._log(f"❌ [{label}] Lỗi: {result.get('message', 'Unknown error')}")
                    success = False

                if result.get("errors"):
                    for err in result["errors"]:
                        self._log(f"   ⚠️ {err}")

            # Upload khuon data separately
            if khuon_entries:
                self._log(f"🔩 [{label}] Đang gửi khuôn ({len(khuon_entries)} khuôn)...")
                if is_remote:
                    khuon_result = upload_khuon_batched(
                        server, username, password, khuon_entries,
                        log_fn=self._log)
                else:
                    khuon_result = upload_khuon(server, username, password, khuon_entries)

                if khuon_result.get("status") in ("ok", "partial"):
                    self._log(f"   ✅ [{label}] Gửi khuôn thành công! {khuon_result.get('khuon_count', 0)} bản ghi")
                    if khuon_result.get("status") == "partial":
                        success = False
                else:
                    self._log(f"   ❌ [{label}] Lỗi khuôn: {khuon_result.get('message', 'Unknown')}")
                    success = False

                if khuon_result.get("errors"):
                    for err in khuon_result["errors"]:
                        self._log(f"   ⚠️ {err}")

        except requests.exceptions.ConnectionError:
            self._log(f"❌ [{label}] Không kết nối được đến server!")
            self._log(f"   Kiểm tra lại URL: {server}")
            success = False
        except Exception as e:
            self._log(f"❌ [{label}] Lỗi: {str(e)}")
            success = False

        return success

    def _do_upload(self):
        try:
            server = self.server_var.get().strip()
            local = self.local_var.get().strip()
            username = self.user_var.get().strip().lower()
            password = self.pass_var.get().strip()
            folder = self.folder_var.get().strip()

            if not all([server, username, password, folder]):
                self._log("❌ Vui lòng điền đầy đủ thông tin!")
                return

            if not os.path.isdir(folder):
                self._log(f"❌ Folder không tồn tại: {folder}")
                return

            # Save config (including password for auto mode)
            self.config["server_url"] = server
            self.config["local_url"] = local
            self.config["username"] = username
            self.config["password"] = password
            self.config["folder"] = folder
            save_config(self.config)

            # Read Excel files
            self._log(f"📂 Đang đọc file Excel từ: {folder}")
            self._log(f"👤 User: {username}")
            self._log(f"🌐 Server chính: {server}")
            if local:
                self._log(f"🏠 Server local: {local}")
            self._log(f"📋 Dây chuyền: MIXER + PL1~PL7")
            self._log("")

            entries, nvvh_entries, loss_entries, khuon_entries, file_results = read_all_files(folder, username)

            for msg in file_results:
                self._log(msg)

            if not entries and not nvvh_entries and not loss_entries and not khuon_entries:
                self._log("\n❌ Không tìm thấy dữ liệu phù hợp!")
                return

            self._log(f"\n📊 Tổng đọc được: {len(entries)} sản lượng, {len(nvvh_entries)} NVVH, {len(loss_entries)} LOSS, {len(khuon_entries)} KHUÔN")

            # ── REMOTE: only send 3 most recent days ──
            RECENT_DAYS = 3
            remote_entries, remote_nvvh, remote_loss, recent_dates = _filter_recent_days(
                entries, nvvh_entries, loss_entries, recent_days=RECENT_DAYS)
            today = date.today()
            day_list = sorted(recent_dates, key=lambda x: (x[0], x[1], x[2]))
            day_strs = [f"{d}/{m}" for (y, m, d) in day_list]
            self._log(f"\n📡 REMOTE chỉ gửi {RECENT_DAYS} ngày gần nhất: {', '.join(day_strs)}")
            self._log(f"   → {len(remote_entries)} SL, {len(remote_nvvh)} NVVH, {len(remote_loss)} LOSS")

            all_ok = True

            # ── 1. Send to REMOTE (3 recent days only) ──
            ok = self._upload_to_server("REMOTE", server, username, password,
                                        remote_entries, remote_nvvh, remote_loss, khuon_entries)
            if not ok:
                all_ok = False

            # ── 2. Start local server + Send to LOCAL (all 31 days) ──
            if local:
                self._start_local_server()
                ok = self._upload_to_server("LOCAL", local, username, password,
                                            entries, nvvh_entries, loss_entries, khuon_entries)
                if not ok:
                    all_ok = False

            # Summary
            self._log(f"\n{'═'*50}")
            if all_ok:
                self._log("🎉 Hoàn tất — tất cả server đều thành công!")
            else:
                self._log("⚠️ Hoàn tất — có server bị lỗi, xem chi tiết ở trên.")

        except Exception as e:
            self._log(f"\n❌ Lỗi: {str(e)}")
        finally:
            self.root.after(0, lambda: self.btn_upload.config(
                state="normal", text="🚀  Gửi tất cả dữ liệu lên hệ thống"))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = UploaderApp()
    app.run()
