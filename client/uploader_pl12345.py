"""
Client Uploader PL1-5 — Gửi dữ liệu sản lượng PL1~PL5, NVVH, LOSS lên server.
Chỉ gửi dữ liệu của PL1, PL2, PL3, PL4, PL5 lên https://sanluongcv.onrender.com/

Sử dụng:
    python uploader_pl12345.py
"""

import os
import sys
import re
import json
import math
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import threading

try:
    import pandas as pd
except ImportError:
    print("Cần cài pandas: pip install pandas openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Cần cài requests: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Cần cài openpyxl: pip install openpyxl")
    sys.exit(1)


# ─── Config ────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    _app_dir = os.path.dirname(sys.executable)
else:
    _app_dir = os.path.dirname(os.path.abspath(__file__))

_profile = ""
for i, arg in enumerate(sys.argv):
    if arg == "--profile" and i + 1 < len(sys.argv):
        _profile = sys.argv[i + 1]

if _profile:
    CONFIG_FILE = os.path.join(_app_dir, f"config_pl12345_{_profile}.json")
else:
    CONFIG_FILE = os.path.join(_app_dir, "config_pl12345.json")

DEFAULT_CONFIG = {
    "server_url": "https://sanluongcv.onrender.com",
    "username": "",
    "password": "",
    "folder": "",
}

# Only PL file type — no MIXER
FILE_CONFIGS = {
    "PL": {
        "pattern": "PL*.xlsx",
        "regex": r"(PL\d+)\s+(\d+)\.(\d+)\.xlsx",
        "sheet_name": "SAN LUONG (2)",
        "usecols": "A:E",
        "skiprows": 8,
        "category": "Pellet Mill",
    },
}

ALL_FILE_TYPES = ["PL"]
# Only PL1 through PL5
ALL_LINES = ["PL1", "PL2", "PL3", "PL4", "PL5"]


# ─── Config Load/Save ─────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
    return DEFAULT_CONFIG.copy()


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ─── Helpers ───────────────────────────────────────────────
def safe_float(val, default=0.0):
    """Convert value to float, replacing NaN/inf with default."""
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


def is_pl1_to_5(line_name):
    """Check if line_name is PL1..PL5."""
    m = re.match(r'^PL(\d+)$', line_name, re.IGNORECASE)
    if m:
        return 1 <= int(m.group(1)) <= 5
    return False


# ─── Excel Parser ──────────────────────────────────────────
def parse_excel_file(filepath, config):
    """Parse a single Excel file. Returns list of day entries."""
    filename = os.path.basename(filepath)
    match = re.match(config["regex"], filename, re.IGNORECASE)
    if not match:
        return None, None

    line_name = match.group(1).upper()

    # *** Filter: only PL1-PL5 ***
    if not is_pl1_to_5(line_name):
        return None, None

    month = int(match.group(2))
    year = int(match.group(3))

    try:
        df = pd.read_excel(
            filepath,
            sheet_name=config["sheet_name"],
            header=None,
            usecols=config["usecols"],
            skiprows=config["skiprows"],
            nrows=31,
        )
        df.columns = ["day", "ca1", "ca2", "ca3", "total"]
        df = df.fillna(0)

        entries = []
        for _, row in df.iterrows():
            day_num = int(row["day"]) if row["day"] else 0
            if day_num < 1 or day_num > 31:
                continue
            entry = {
                "line_name": line_name,
                "category": config["category"],
                "year": year,
                "month": month,
                "day": day_num,
                "ca1": round(safe_float(row["ca1"]), 2),
                "ca2": round(safe_float(row["ca2"]), 2),
                "ca3": round(safe_float(row["ca3"]), 2),
                "total": round(safe_float(row["total"]), 2),
                "cam_bot": 0,
            }
            entries.append(entry)

        return f"{line_name} T{month}.{year}", entries

    except Exception as e:
        return filename, f"Lỗi: {e}"


# ─── NVVH Parser ───────────────────────────────────────────
def parse_nvvh_from_file(filepath, month, year):
    """Parse NVVH (operator names) from PL1 file only (for pl1_5 group)."""
    filename = os.path.basename(filepath).upper()

    # Only process PL1-PL5 files
    if not any(f"PL{i}" in filename for i in range(1, 6)):
        return []

    pl_group = "pl1_5"

    # Only read NVVH from PL1
    if "PL1 " not in filename and "PL1." not in filename:
        return []

    nvvh_row = 49
    ca_cols = {"ca1": 33, "ca2": 37, "ca3": 40}

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for day in range(1, 32):
            sheet_name = str(day)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=nvvh_row, max_row=nvvh_row, values_only=False):
                for ca_key, col_idx in ca_cols.items():
                    if col_idx < len(row):
                        raw = row[col_idx].value
                        if raw:
                            names_str = str(raw).strip()
                            parts = re.split(r'[+\-]', names_str)
                            names_str = ','.join(n.strip() for n in parts if n.strip())
                            entries.append({
                                "year": year, "month": month,
                                "day": day, "ca": ca_key,
                                "pl_group": pl_group, "names": names_str,
                            })
        wb.close()
    except Exception as e:
        print(f"⚠️ NVVH parse error ({filepath}): {e}")

    return entries


# ─── LOSS Parser ───────────────────────────────────────────
def parse_loss_from_file(filepath, month, year):
    """Parse LOSS sheet from a PL1-PL5 file."""
    filename = os.path.basename(filepath).upper()

    # Only PL files
    pl_match = re.search(r'PL(\d+)', filename)
    if not pl_match:
        return []
    pl_num = int(pl_match.group(1))

    # *** Filter: only PL1-PL5 ***
    if pl_num < 1 or pl_num > 5:
        return []

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "LOSS" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["LOSS"]
        all_rows = list(ws.iter_rows(min_row=7, max_row=800, values_only=False))
        wb.close()

        for day in range(1, 32):
            start_col = 5 + (day - 1) * 6
            for row in all_rows:
                code_cell = row[1] if len(row) > 1 else None
                desc_cell = row[2] if len(row) > 2 else None
                code = code_cell.value if code_cell else None
                if code is None:
                    continue

                desc = str(desc_cell.value) if desc_cell and desc_cell.value else ""

                vals = []
                for offset in range(6):
                    col_idx = start_col - 1 + offset
                    v = row[col_idx].value if col_idx < len(row) else None
                    try:
                        vals.append(float(v) if v else 0)
                    except (ValueError, TypeError):
                        vals.append(0)

                if any(v > 0 for v in vals):
                    # PL time is in hours → convert to minutes
                    time_mult = 60
                    entries.append({
                        "year": year, "month": month,
                        "day": day, "pl_num": pl_num,
                        "code": str(code), "description": desc,
                        "ca1_count": int(vals[0]),
                        "ca1_time": round(vals[1] * time_mult),
                        "ca2_count": int(vals[2]),
                        "ca2_time": round(vals[3] * time_mult),
                        "ca3_count": int(vals[4]),
                        "ca3_time": round(vals[5] * time_mult),
                    })
    except Exception as e:
        print(f"⚠️ LOSS parse error ({filepath}): {e}")

    return entries


# ─── Khuôn Parser ─────────────────────────────────────────
def parse_khuon_from_file(filepath, pl_num, month, year):
    """Parse SERI sheet from a PL file for khuon tracking data."""
    # *** Filter: only PL1-PL5 ***
    if pl_num < 1 or pl_num > 5:
        return []

    if pl_num == 2:
        seri_idx, khuon_idx, day1_idx = 2, 3, 4
    else:
        seri_idx, khuon_idx, day1_idx = 1, 2, 3

    tong_thang_idx = day1_idx + 31
    ton_truoc_idx = day1_idx + 32
    tong_idx = day1_idx + 33

    entries = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if 'SERI' not in wb.sheetnames:
            wb.close()
            return []

        ws = wb['SERI']
        for row in ws.iter_rows(min_row=3, max_row=100, min_col=1,
                                max_col=tong_idx + 1, values_only=True):
            seri_val = row[seri_idx] if len(row) > seri_idx else None
            if seri_val is None or str(seri_val).strip() == '':
                continue

            thong_so = row[khuon_idx] if len(row) > khuon_idx else ''

            days = {}
            for d in range(31):
                col_idx = day1_idx + d
                val = row[col_idx] if len(row) > col_idx else 0
                v = round(float(val), 2) if val else 0
                if v > 0:
                    days[d + 1] = v

            tong_thang = row[tong_thang_idx] if len(row) > tong_thang_idx else 0
            ton_truoc = row[ton_truoc_idx] if len(row) > ton_truoc_idx else 0
            tong = row[tong_idx] if len(row) > tong_idx else 0

            entries.append({
                'pl_num': pl_num,
                'year': year,
                'month': month,
                'seri': str(seri_val).strip(),
                'thong_so': str(thong_so).strip() if thong_so else '',
                'days': days,
                'tong_thang': round(float(tong_thang), 2) if tong_thang else 0,
                'ton_truoc': round(float(ton_truoc), 2) if ton_truoc else 0,
                'tong': round(float(tong), 2) if tong else 0,
            })

        wb.close()
    except Exception as e:
        print(f"⚠️ Khuôn parse error ({filepath}): {e}")

    return entries


def read_all_files(folder, username):
    """Read PL1-PL5 Excel files from folder.
    Returns: (production_entries, nvvh_entries, loss_entries, khuon_entries, file_results)
    """
    all_entries = []
    nvvh_entries = []
    loss_entries = []
    khuon_entries = []
    file_results = []

    for type_key in ALL_FILE_TYPES:
        config = FILE_CONFIGS[type_key]
        pattern = os.path.join(folder, config["pattern"])
        files = sorted(glob.glob(pattern))

        for filepath in files:
            label, result = parse_excel_file(filepath, config)
            if result is None:
                continue

            if isinstance(result, str):
                file_results.append(f"❌ {label}: {result}")
                continue

            if result:
                all_entries.extend(result)
                total = sum(e["total"] for e in result)
                file_results.append(f"✅ {label}: {len(result)} ngày, {total:,.1f} tấn")

            # Parse NVVH + LOSS
            fn = os.path.basename(filepath)
            match = re.match(config["regex"], fn, re.IGNORECASE)
            if match:
                month = int(match.group(2))
                year = int(match.group(3))
                nvvh = parse_nvvh_from_file(filepath, month, year)
                if nvvh:
                    nvvh_entries.extend(nvvh)
                    file_results.append(f"   👤 NVVH: {len(nvvh)} records")
                loss = parse_loss_from_file(filepath, month, year)
                if loss:
                    loss_entries.extend(loss)
                    file_results.append(f"   📋 LOSS: {len(loss)} records")

            # Parse KHUÔN
            if match:
                line_name = match.group(1).upper()
                pl_num = int(re.search(r'\d+', line_name).group())
                month = int(match.group(2))
                year = int(match.group(3))
                khuon = parse_khuon_from_file(filepath, pl_num, month, year)
                if khuon:
                    khuon_entries.extend(khuon)
                    file_results.append(f"   🔩 KHUÔN: {len(khuon)} khuôn")

    return all_entries, nvvh_entries, loss_entries, khuon_entries, file_results


def upload_data(server_url, username, password, entries, nvvh_entries=None, loss_entries=None):
    """Send production, NVVH, and LOSS data to server via API."""
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    url = f"{server_url.rstrip('/')}/api/upload-data"
    payload = {
        "username": username,
        "password": password,
        "entries": entries,
        "nvvh_entries": nvvh_entries or [],
        "loss_entries": loss_entries or [],
    }
    resp = requests.post(url, json=payload, timeout=300, verify=False)
    try:
        return resp.json()
    except (json.JSONDecodeError, ValueError):
        body_preview = resp.text[:200].strip() if resp.text else "(empty)"
        return {
            "status": "error",
            "message": f"Server trả về HTTP {resp.status_code} — không phải JSON. Nội dung: {body_preview}",
        }


def upload_khuon(server_url, username, password, khuon_entries):
    """Send khuon tracking data to server via API."""
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    url = f"{server_url.rstrip('/')}/api/upload-khuon"
    payload = {
        "username": username,
        "password": password,
        "khuon_entries": khuon_entries,
    }
    resp = requests.post(url, json=payload, timeout=300, verify=False)
    try:
        return resp.json()
    except (json.JSONDecodeError, ValueError):
        body_preview = resp.text[:200].strip() if resp.text else "(empty)"
        return {
            "status": "error",
            "message": f"Server trả về HTTP {resp.status_code} — không phải JSON. Nội dung: {body_preview}",
        }


# ─── GUI ───────────────────────────────────────────────────
class UploaderApp:
    def __init__(self):
        self.config = load_config()
        self.root = tk.Tk()
        title = "📊 Cập Nhật Sản Lượng — PL1~PL5"
        if _profile:
            title += f" — {_profile.upper()}"
        self.root.title(title)
        self.root.geometry("520x520")
        self.root.resizable(False, False)
        self.root.configure(bg="#1a1a2e")
        self._build_ui()

    def _build_ui(self):
        bg = "#1a1a2e"
        fg = "#e2e8f0"
        accent = "#6366f1"
        entry_bg = "#16213e"

        tk.Label(self.root, text="📊 Cập Nhật Sản Lượng PL1~PL5",
                 font=("Segoe UI", 16, "bold"), bg=bg, fg=fg).pack(pady=(15, 5))
        tk.Label(self.root, text="Chỉ gửi PL1-PL5 → sanluongcv.onrender.com",
                 font=("Segoe UI", 9), bg=bg, fg="#94a3b8").pack()

        form = tk.Frame(self.root, bg=bg)
        form.pack(pady=15, padx=25, fill="x")

        tk.Label(form, text="Server URL:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=0, column=0, sticky="w", pady=4)
        self.server_var = tk.StringVar(value=self.config["server_url"])
        tk.Entry(form, textvariable=self.server_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=0, column=1, columnspan=2, pady=4, sticky="ew")

        tk.Label(form, text="Username:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=1, column=0, sticky="w", pady=4)
        self.user_var = tk.StringVar(value=self.config["username"])
        tk.Entry(form, textvariable=self.user_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=1, column=1, columnspan=2, pady=4, sticky="ew")

        tk.Label(form, text="Password:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=2, column=0, sticky="w", pady=4)
        self.pass_var = tk.StringVar(value=self.config.get("password", ""))
        tk.Entry(form, textvariable=self.pass_var, font=("Segoe UI", 10),
                 show="•", bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=35).grid(row=2, column=1, columnspan=2, pady=4, sticky="ew")

        tk.Label(form, text="Folder Excel:", font=("Segoe UI", 10),
                 bg=bg, fg=fg).grid(row=3, column=0, sticky="w", pady=4)
        self.folder_var = tk.StringVar(value=self.config["folder"])
        tk.Entry(form, textvariable=self.folder_var, font=("Segoe UI", 10),
                 bg=entry_bg, fg=fg, insertbackground=fg, relief="flat",
                 width=28).grid(row=3, column=1, pady=4, sticky="ew")
        tk.Button(form, text="📁", font=("Segoe UI", 10),
                  bg=accent, fg="white", relief="flat", width=3,
                  command=self._browse_folder).grid(row=3, column=2, padx=(5, 0))

        form.columnconfigure(1, weight=1)

        self.btn_upload = tk.Button(
            self.root, text="🚀  Gửi PL1~PL5 lên hệ thống",
            font=("Segoe UI", 12, "bold"), bg=accent, fg="white",
            activebackground="#4f46e5", relief="flat", height=2,
            command=self._start_upload)
        self.btn_upload.pack(pady=10, padx=25, fill="x")

        self.status_text = tk.Text(
            self.root, height=10, bg="#0f172a", fg="#94a3b8",
            font=("Consolas", 9), relief="flat", state="disabled",
            wrap="word")
        self.status_text.pack(padx=25, pady=(0, 15), fill="both", expand=True)

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa file Excel PL1-PL5")
        if folder:
            self.folder_var.set(folder)

    def _log(self, msg):
        self.status_text.config(state="normal")
        self.status_text.insert("end", msg + "\n")
        self.status_text.see("end")
        self.status_text.config(state="disabled")

    def _clear_log(self):
        self.status_text.config(state="normal")
        self.status_text.delete("1.0", "end")
        self.status_text.config(state="disabled")

    def _start_upload(self):
        self.btn_upload.config(state="disabled", text="⏳ Đang xử lý...")
        self._clear_log()
        threading.Thread(target=self._do_upload, daemon=True).start()

    def _do_upload(self):
        try:
            server = self.server_var.get().strip()
            username = self.user_var.get().strip().lower()
            password = self.pass_var.get().strip()
            folder = self.folder_var.get().strip()

            if not all([server, username, password, folder]):
                self._log("❌ Vui lòng điền đầy đủ thông tin!")
                return

            if not os.path.isdir(folder):
                self._log(f"❌ Folder không tồn tại: {folder}")
                return

            self.config["server_url"] = server
            self.config["username"] = username
            self.config["password"] = password
            self.config["folder"] = folder
            save_config(self.config)

            self._log(f"📂 Đang đọc file PL1~PL5 từ: {folder}")
            self._log(f"👤 User: {username}")
            self._log(f"🌐 Server: {server}")
            self._log("")

            entries, nvvh_entries, loss_entries, khuon_entries, file_results = read_all_files(folder, username)

            for msg in file_results:
                self._log(msg)

            if not entries and not nvvh_entries and not loss_entries and not khuon_entries:
                self._log("\n❌ Không tìm thấy dữ liệu PL1~PL5 phù hợp!")
                return

            self._log(f"\n📊 Tổng: {len(entries)} sản lượng, {len(nvvh_entries)} NVVH, {len(loss_entries)} LOSS, {len(khuon_entries)} KHUÔN")
            self._log(f"🚀 Đang gửi lên {server}...")

            if entries or nvvh_entries or loss_entries:
                result = upload_data(server, username, password, entries, nvvh_entries, loss_entries)

                if result.get("status") == "ok":
                    now = datetime.now().strftime("%H:%M:%S")
                    self._log(f"\n✅ Gửi sản lượng thành công lúc {now}!")
                    self._log(f"   Sản lượng: {result.get('inserted', 0)} bản ghi")
                    if result.get('nvvh_count', 0) > 0:
                        self._log(f"   NVVH: {result['nvvh_count']} bản ghi")
                    if result.get('loss_count', 0) > 0:
                        self._log(f"   LOSS: {result['loss_count']} bản ghi")
                    if result.get("skipped", 0) > 0:
                        self._log(f"   Bỏ qua: {result['skipped']}")
                else:
                    self._log(f"\n❌ Lỗi: {result.get('message', 'Unknown error')}")

                if result.get("errors"):
                    for err in result["errors"]:
                        self._log(f"   ⚠️ {err}")

            if khuon_entries:
                self._log(f"\n🔩 Đang gửi dữ liệu khuôn ({len(khuon_entries)} khuôn)...")
                khuon_result = upload_khuon(server, username, password, khuon_entries)

                if khuon_result.get("status") == "ok":
                    self._log(f"   ✅ Gửi khuôn thành công! {khuon_result.get('khuon_count', 0)} bản ghi")
                else:
                    self._log(f"   ❌ Lỗi khuôn: {khuon_result.get('message', 'Unknown')}")

                if khuon_result.get("errors"):
                    for err in khuon_result["errors"]:
                        self._log(f"   ⚠️ {err}")

        except requests.exceptions.ConnectionError:
            self._log(f"\n❌ Không kết nối được đến server!")
            self._log(f"   Kiểm tra lại URL: {self.server_var.get()}")
        except Exception as e:
            self._log(f"\n❌ Lỗi: {str(e)}")
        finally:
            self.root.after(0, lambda: self.btn_upload.config(
                state="normal", text="🚀  Gửi PL1~PL5 lên hệ thống"))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = UploaderApp()
    app.run()
