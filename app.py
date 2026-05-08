"""
Flask Web App - Sản Lượng Hàng Ngày
Displays daily production output from SQLite database.
Data imported from Excel files (PL1-PL7, MIXER) into production.db.
Accessible via LAN (company WiFi) and Internet.
"""

import os
import time
from functools import wraps
from flask import (Flask, render_template, jsonify, request,
                   session, redirect, url_for, send_file)
from database import (init_db, import_from_excel, import_month_from_excel,
                      load_all_from_db, get_db_stats,
                      save_manual_input, get_manual_inputs,
                      get_monthly_sale_total, save_uploaded_data,
                      authenticate, change_password, get_all_users,
                      save_khuon_data, get_khuon_data, get_khuon_yearly)

app = Flask(__name__)
app.secret_key = "cpvn-sanluong-2026-secret"

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# Initialize database on startup
init_db(DATA_DIR)


@app.route("/healthz")
def healthz():
    """Health check endpoint for cloud platforms (Render, etc.)."""
    return "ok", 200

# ── In-memory cache (reads from SQLite, not Excel) ───────────
CACHE_TTL = 300  # 5 minutes
_cache = {"data": None, "timestamp": 0}


def get_all_data():
    """Return cached data, refreshing from SQLite only when stale."""
    now = time.time()
    if _cache["data"] is None or (now - _cache["timestamp"]) > CACHE_TTL:
        _cache["data"] = load_all_from_db(DATA_DIR)
        _cache["timestamp"] = now
    return _cache["data"]


def invalidate_cache():
    """Force reload on next request."""
    _cache["data"] = None
    _cache["timestamp"] = 0


# ── Authentication ───────────────────────────────────────────

def login_required(f):
    """Decorator: redirect to /login if not authenticated."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page."""
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "").strip()
        user = authenticate(DATA_DIR, username, password)
        if user:
            session["user"] = user
            return redirect(url_for("index"))
        return render_template("login.html", error="Sai tài khoản hoặc mật khẩu!")
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    """Logout and redirect to login page."""
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/api/lib/h2c")
def serve_h2c():
    """Serve html2canvas library dynamically to bypass antivirus blocking."""
    h2c_path = os.path.join(DATA_DIR, "node_modules", "html2canvas", "dist", "html2canvas.min.js")
    if not os.path.exists(h2c_path):
        h2c_path = os.path.join(DATA_DIR, "static", "h2c_lib.js")
    try:
        with open(h2c_path, "r", encoding="utf-8") as f:
            content = f.read()
        return app.response_class(content, mimetype="application/javascript")
    except Exception:
        return "// html2canvas not found", 404


@app.route("/api/upload-data", methods=["POST"])
def api_upload_data():
    """Receive production, NVVH, and LOSS data from client PCs.
    JSON body: {username, password, entries: [...], nvvh_entries: [...], loss_entries: [...]}
    """
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "No JSON data"}), 400

    username = data.get("username", "").strip().lower()
    password = data.get("password", "").strip()
    entries = data.get("entries", [])
    nvvh_entries = data.get("nvvh_entries", [])
    loss_entries = data.get("loss_entries", [])

    # Authenticate
    user = authenticate(DATA_DIR, username, password)
    if not user:
        return jsonify({"status": "error", "message": "Sai tài khoản hoặc mật khẩu"}), 401

    if not entries and not nvvh_entries and not loss_entries:
        return jsonify({"status": "error", "message": "Không có dữ liệu"}), 400

    # Save data
    result = save_uploaded_data(DATA_DIR, username, entries, nvvh_entries, loss_entries)

    # Invalidate cache so dashboard shows new data
    invalidate_cache()

    return jsonify(result)


@app.route("/api/upload-khuon", methods=["POST"])
def api_upload_khuon():
    """Receive khuon tracking data from client uploader.
    JSON body: {username, password, khuon_entries: [...]}
    """
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "No JSON data"}), 400

    username = data.get("username", "").strip().lower()
    password = data.get("password", "").strip()
    khuon_entries = data.get("khuon_entries", [])

    # Authenticate
    user = authenticate(DATA_DIR, username, password)
    if not user:
        return jsonify({"status": "error", "message": "Sai tài khoản hoặc mật khẩu"}), 401

    if not khuon_entries:
        return jsonify({"status": "error", "message": "Không có dữ liệu khuôn"}), 400

    result = save_khuon_data(DATA_DIR, username, khuon_entries)
    return jsonify(result)


@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    """Change password for the current user."""
    data = request.get_json()
    new_pass = data.get("new_password", "").strip()
    if not new_pass or len(new_pass) < 1:
        return jsonify({"error": "Mật khẩu không được để trống"}), 400

    username = session["user"]["username"]
    if change_password(DATA_DIR, username, new_pass):
        return jsonify({"status": "ok", "message": "Đổi mật khẩu thành công!"})
    return jsonify({"error": "Lỗi đổi mật khẩu"}), 500


@app.route("/api/users")
@login_required
def api_users():
    """Admin only: get all users with passwords."""
    if session["user"]["role"] != "admin":
        return jsonify({"error": "Không có quyền"}), 403
    users = get_all_users(DATA_DIR)
    return jsonify(users)


@app.context_processor
def inject_user():
    """Make current_user available in all templates."""
    return {"current_user": session.get("user", None)}


@app.route("/")
@login_required
def index():
    result = get_all_data()
    months = result["months"]  # sorted descending (newest first)
    all_data = result["data"]

    # Default to latest month, or use query param, or current month
    from datetime import datetime as _dt
    fallback = _dt.now().strftime("%Y-%m")
    selected_month = request.args.get("month", months[0] if months else fallback)

    # Get data for selected month
    month_data = all_data.get(selected_month, {})

    # Build month labels for dropdown (e.g., "Tháng 1/2026")
    month_labels = []
    for mk in months:
        parts = mk.split("-")
        label = f"Tháng {int(parts[1])}/{parts[0]}"
        month_labels.append({"key": mk, "label": label})

    # Monthly sale total for dashboard
    sale_total = 0
    if selected_month and "-" in selected_month:
        parts = selected_month.split("-")
        sale_total = get_monthly_sale_total(DATA_DIR, int(parts[0]), int(parts[1]))

    return render_template(
        "index.html",
        data=month_data,
        months=month_labels,
        selected_month=selected_month,
        monthly_sale_total=sale_total,
    )


@app.route("/api/data")
@login_required
def api_all_data():
    """JSON API: data for a specific month (or latest)."""
    result = get_all_data()
    months = result["months"]
    from datetime import datetime as _dt
    fallback = _dt.now().strftime("%Y-%m")
    selected = request.args.get("month", months[0] if months else fallback)
    month_data = result["data"].get(selected, {})

    # Monthly sale total
    sale_total = 0
    if selected and "-" in selected:
        parts = selected.split("-")
        sale_total = get_monthly_sale_total(DATA_DIR, int(parts[0]), int(parts[1]))

    return jsonify({
        "months": [
            {"key": mk, "label": f"Tháng {int(mk.split('-')[1])}/{mk.split('-')[0]}"}
            for mk in months
        ],
        "selected_month": selected,
        "lines": month_data,
        "monthly_sale_total": sale_total,
    })


@app.route("/api/data/<line>")
@login_required
def api_line_data(line):
    """JSON API: data for a specific production line in the selected month."""
    result = get_all_data()
    months = result["months"]
    selected = request.args.get("month", months[0] if months else "")
    month_data = result["data"].get(selected, {})

    line_upper = line.upper()
    if line_upper in month_data:
        return jsonify(month_data[line_upper])
    return jsonify({"error": f"Line {line} not found in {selected}"}), 404


@app.route("/api/refresh", methods=["POST"])
@login_required
def api_refresh():
    """Re-import Excel → SQLite for the selected month (or all)."""
    month = request.args.get("month", "")
    if month:
        # Only re-import the selected month (fast)
        count = import_month_from_excel(DATA_DIR, month)
    else:
        # Full re-import (all months)
        count = import_from_excel(DATA_DIR)
    invalidate_cache()
    return jsonify({"status": "ok", "records": count, "month": month})


@app.route("/api/db-stats")
def api_db_stats():
    """Database statistics."""
    stats = get_db_stats(DATA_DIR)
    return jsonify(stats or {"error": "No database found"})


@app.route("/api/manual-input", methods=["POST"])
@login_required
def api_save_manual_input():
    """Save SALE or STOCK value for a specific day."""
    data = request.get_json()
    month_key = data.get("month", "")
    day = data.get("day", 0)
    field = data.get("field", "")  # 'sale' or 'stock'
    value = data.get("value", 0)

    if not month_key or not day or field not in ("sale", "stock"):
        return jsonify({"error": "Invalid input"}), 400

    parts = month_key.split("-")
    year = int(parts[0])
    month = int(parts[1])

    save_manual_input(DATA_DIR, year, month, day, field, float(value))
    return jsonify({"status": "ok", "field": field, "day": day, "value": value})


@app.route("/report")
@login_required
def report_page():
    """Daily report page with 31 day buttons."""
    result = get_all_data()
    months = result["months"]
    from datetime import datetime as _dt
    fallback = _dt.now().strftime("%Y-%m")
    selected_month = request.args.get("month", months[0] if months else fallback)
    month_data = result["data"].get(selected_month, {})

    # Build month labels
    month_labels = []
    for mk in months:
        parts = mk.split("-")
        label = f"Tháng {int(parts[1])}/{parts[0]}"
        month_labels.append({"key": mk, "label": label})

    # Find days that have data (any line with total > 0)
    days_with_data = set()
    for line_info in month_data.values():
        for d in line_info.get("days", []):
            if d["total"] > 0:
                days_with_data.add(d["day"])
    days_with_data = sorted(days_with_data)

    # Read html2canvas for inline embedding (avoids antivirus blocking)
    h2c_path = os.path.join(DATA_DIR, "static", "html2canvas.min.js")
    html2canvas_js = ""
    if os.path.exists(h2c_path):
        with open(h2c_path, "r", encoding="utf-8") as f:
            html2canvas_js = f.read()

    return render_template(
        "report.html",
        data=month_data,
        months=month_labels,
        selected_month=selected_month,
        days_with_data=days_with_data,
        html2canvas_js=html2canvas_js,
    )


@app.route("/api/report/<int:day>")
@login_required
def api_report_day(day):
    """JSON API: daily report for a specific day."""
    result = get_all_data()
    months = result["months"]
    selected = request.args.get("month", months[0] if months else "")
    month_data = result["data"].get(selected, {})

    # Build report
    report = {"day": day, "month": selected}

    # MIXER data
    mixer = month_data.get("MIXER", {})
    mixer_day = next((d for d in mixer.get("days", []) if d["day"] == day), None)
    report["mixer"] = {
        "ca1": mixer_day["ca1"] if mixer_day else 0,
        "ca2": mixer_day["ca2"] if mixer_day else 0,
        "ca3": mixer_day["ca3"] if mixer_day else 0,
        "total": mixer_day["total"] if mixer_day else 0,
        "cam_bot": mixer_day.get("cam_bot", 0) if mixer_day else 0,
    }

    # PL1-PL7 data
    pellets = []
    for i in range(1, 8):
        pl_name = f"PL{i}"
        pl = month_data.get(pl_name, {})
        pl_day = next((d for d in pl.get("days", []) if d["day"] == day), None)
        pellets.append({
            "name": pl_name,
            "ca1": pl_day["ca1"] if pl_day else 0,
            "ca2": pl_day["ca2"] if pl_day else 0,
            "ca3": pl_day["ca3"] if pl_day else 0,
            "total": pl_day["total"] if pl_day else 0,
        })
    report["pellets"] = pellets

    # Total pellet
    report["total_pellet"] = {
        "ca1": round(sum(p["ca1"] for p in pellets), 2),
        "ca2": round(sum(p["ca2"] for p in pellets), 2),
        "ca3": round(sum(p["ca3"] for p in pellets), 2),
        "total": round(sum(p["total"] for p in pellets), 2),
    }

    # Manual inputs (sale/stock)
    parts = selected.split("-")
    manual = get_manual_inputs(DATA_DIR, int(parts[0]), int(parts[1]), day)
    report["sale"] = manual.get("sale", None)
    report["stock"] = manual.get("stock", None)

    return jsonify(report)


@app.route("/api/report/<int:day>/details")
@login_required
def api_report_day_details(day):
    """JSON API: NVVH + LOSS details (from database — fast)."""
    result = get_all_data()
    months = result["months"]
    selected = request.args.get("month", months[0] if months else "")
    parts = selected.split("-")
    year_num = int(parts[0])
    month_num = int(parts[1])

    from database import get_nvvh_for_day, get_loss_for_day
    details = {
        "nvvh": get_nvvh_for_day(year_num, month_num, day),
        "loss_notes": get_loss_for_day(year_num, month_num, day),
    }

    return jsonify(details)


# ── Khuôn Tracking ────────────────────────────────────────

def _load_khuon_for_month(year, month):
    """Load khuon data: try DB first, fall back to Excel."""
    # Try DB first
    khuon_data = get_khuon_data(DATA_DIR, year, month)
    total_in_db = sum(len(v) for v in khuon_data.values())
    if total_in_db > 0:
        return khuon_data
    # Fall back to Excel (local dev)
    try:
        from khuon_loader import load_all_khuon_for_month
        return load_all_khuon_for_month(DATA_DIR, month, year)
    except Exception:
        return khuon_data


def _load_khuon_yearly_data(year):
    """Load khuon yearly: try DB first, fall back to Excel."""
    data = get_khuon_yearly(DATA_DIR, year)
    total_in_db = sum(len(v) for v in data.values())
    if total_in_db > 0:
        return data
    try:
        from khuon_loader import load_khuon_yearly
        return load_khuon_yearly(DATA_DIR, year)
    except Exception:
        return data


@app.route("/khuon")
@login_required
def khuon_page():
    """Mold tracking page — daily & monthly views."""
    result = get_all_data()
    months = result["months"]
    from datetime import datetime as _dt
    fallback = _dt.now().strftime("%Y-%m")
    selected_month = request.args.get("month", months[0] if months else fallback)

    # Parse month
    parts = selected_month.split("-")
    year = int(parts[0])
    month = int(parts[1])
    month_label = f"Tháng {month}/{year}"

    # Build month labels for dropdown
    month_labels = []
    for mk in months:
        p = mk.split("-")
        label = f"Tháng {int(p[1])}/{p[0]}"
        month_labels.append({"key": mk, "label": label})

    # Load khuon data (DB first, then Excel fallback)
    khuon_data = _load_khuon_for_month(year, month)

    return render_template(
        "khuon.html",
        khuon_data=khuon_data,
        months=month_labels,
        selected_month=selected_month,
        month_label=month_label,
    )


@app.route("/api/khuon")
@login_required
def api_khuon_monthly():
    """API: khuon data for a specific month."""
    month_key = request.args.get("month", "")
    if not month_key or "-" not in month_key:
        return jsonify({"error": "Missing month param"}), 400

    parts = month_key.split("-")
    year = int(parts[0])
    month = int(parts[1])
    data = _load_khuon_for_month(year, month)
    return jsonify(data)


@app.route("/api/khuon/yearly")
@login_required
def api_khuon_yearly():
    """API: yearly khuon summary (all months in a year)."""
    year = request.args.get("year", "")
    if not year:
        from datetime import datetime as _dt
        year = _dt.now().year
    else:
        year = int(year)

    data = _load_khuon_yearly_data(year)
    return jsonify(data)


@app.route("/api/khuon/export")
@login_required
def api_khuon_export():
    """Export khuon data to Excel file."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    view = request.args.get("view", "monthly")  # daily or monthly
    month_key = request.args.get("month", "")

    if not month_key or "-" not in month_key:
        from datetime import datetime as _dt
        month_key = _dt.now().strftime("%Y-%m")

    parts = month_key.split("-")
    year = int(parts[0])
    month = int(parts[1])

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid")
    value_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
    summary_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')

    if view == "daily":
        # Daily view
        data = _load_khuon_for_month(year, month)
        ws = wb.active
        ws.title = f"T{month}.{year}"

        # Header
        headers = ["STT", "Seri Khuôn", "Khuôn", "PL"]
        for d in range(1, 32):
            headers.append(str(d))
        headers += ["Tổng Tháng", "Tồn Trước", "Tổng"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Data
        row_num = 2
        for pl_key in [f"PL{i}" for i in range(1, 8)]:
            for m in data.get(pl_key, []):
                ws.cell(row=row_num, column=1, value=row_num - 1).alignment = center
                ws.cell(row=row_num, column=2, value=m["seri"])
                ws.cell(row=row_num, column=3, value=m.get("thong_so", ""))
                ws.cell(row=row_num, column=4, value=pl_key).alignment = center
                for d in range(1, 32):
                    val = m.get("days", {}).get(str(d), m.get("days", {}).get(d, 0)) or 0
                    cell = ws.cell(row=row_num, column=4 + d, value=float(val))
                    if float(val) > 0:
                        cell.fill = value_fill
                    cell.alignment = center
                ws.cell(row=row_num, column=36, value=m.get("tong_thang", 0)).font = summary_font
                ws.cell(row=row_num, column=37, value=m.get("ton_truoc", 0))
                ws.cell(row=row_num, column=38, value=m.get("tong", 0)).font = summary_font
                for col in range(1, 39):
                    ws.cell(row=row_num, column=col).border = thin_border
                row_num += 1

        # Column widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        filename = f"Khuon_T{month}_{year}.xlsx"

    else:
        # Monthly/Yearly view
        data = _load_khuon_yearly_data(year)
        ws = wb.active
        ws.title = f"Nam {year}"

        headers = ["STT", "Seri Khuôn", "Khuôn", "PL"]
        for m in range(1, 13):
            headers.append(f"T{m}")
        headers.append("Tổng Năm")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        row_num = 2
        for pl_key in [f"PL{i}" for i in range(1, 8)]:
            for m in data.get(pl_key, []):
                ws.cell(row=row_num, column=1, value=row_num - 1).alignment = center
                ws.cell(row=row_num, column=2, value=m["seri"])
                ws.cell(row=row_num, column=3, value=m.get("thong_so", ""))
                ws.cell(row=row_num, column=4, value=pl_key).alignment = center
                months_data = m.get("months", {})
                for mo in range(1, 13):
                    val = months_data.get(str(mo), months_data.get(mo, 0)) or 0
                    cell = ws.cell(row=row_num, column=4 + mo, value=float(val))
                    if float(val) > 0:
                        cell.fill = value_fill
                    cell.alignment = center
                ws.cell(row=row_num, column=17, value=m.get("year_total", 0)).font = summary_font
                for col in range(1, 18):
                    ws.cell(row=row_num, column=col).border = thin_border
                row_num += 1

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        filename = f"Khuon_Nam_{year}.xlsx"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

if __name__ == "__main__":
    print("=" * 60)
    print("  SẢN LƯỢNG HÀNG NGÀY - Flask Web App")
    print("  Truy cập: http://localhost:5000")
    print("  Hoặc từ thiết bị khác: http://<IP máy>:5000")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=True)
