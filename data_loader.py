"""
Data loader module - Reads production data from Excel files.
Supports two file types:
  1. PL files (PL1-PL7): sheet "SAN LUONG (2)", cols A:E, rows 9-39
  2. MIXER files: sheet "SAN LUONG", cols B:F, rows 8-38

Returns data grouped by month/year for filtering.
"""

import os
import re
import glob
import pandas as pd


# File type configurations
FILE_CONFIGS = {
    "PL": {
        "pattern": "PL*.xlsx",
        "regex": r"(PL\d+)\s+(\d+)\.(\d+)\.xlsx",
        "sheet_name": "SAN LUONG (2)",
        "usecols": "A:E",
        "skiprows": 8,
        "category": "Pellet Mill",
    },
    "MIXER": {
        "pattern": "MIXER*.xls*",
        "regex": r"(MIXER)\s+T(\d+)\.(\d+)\.(xlsx|xlsm)",
        "sheet_name": "SAN LUONG",
        "usecols": "B:F",
        "skiprows": 7,
        "category": "Mixer",
        "extra_col": "AL",  # Column AL = Sản lượng cám bột
    },
}


def _parse_file(filepath, config):
    """Parse a single Excel file based on its config."""
    filename = os.path.basename(filepath)
    match = re.match(config["regex"], filename, re.IGNORECASE)
    if not match:
        return None

    line_name = match.group(1).upper()
    month = int(match.group(2))
    year = int(match.group(3))

    try:
        df = pd.read_excel(
            filepath,
            sheet_name=config["sheet_name"],
            header=None,
            usecols=config["usecols"],
            skiprows=config["skiprows"],
            nrows=31,
        )
        df.columns = ["day", "ca1", "ca2", "ca3", "total"]
        df = df.fillna(0)

        days = []
        for _, row in df.iterrows():
            day_num = int(row["day"]) if row["day"] else 0
            if day_num < 1 or day_num > 31:
                continue
            days.append({
                "day": day_num,
                "ca1": round(float(row["ca1"]), 2),
                "ca2": round(float(row["ca2"]), 2),
                "ca3": round(float(row["ca3"]), 2),
                "total": round(float(row["total"]), 2),
            })

        summary = {
            "ca1": round(sum(d["ca1"] for d in days), 2),
            "ca2": round(sum(d["ca2"] for d in days), 2),
            "ca3": round(sum(d["ca3"] for d in days), 2),
            "total": round(sum(d["total"] for d in days), 2),
        }

        result = {
            "name": line_name,
            "month": month,
            "year": year,
            "month_key": f"{year}-{month:02d}",
            "category": config["category"],
            "days": days,
            "summary": summary,
        }

        # Read extra column if configured (e.g., MIXER column AL = cám bột)
        extra_col = config.get("extra_col")
        if extra_col:
            try:
                df_extra = pd.read_excel(
                    filepath,
                    sheet_name=config["sheet_name"],
                    header=None,
                    usecols=extra_col,
                    skiprows=config["skiprows"],
                    nrows=31,
                )
                for idx, row in df_extra.iterrows():
                    day_idx = idx  # 0-based
                    if day_idx < len(days):
                        val = row.iloc[0]
                        days[day_idx]["cam_bot"] = round(float(val) if val else 0, 2)
            except Exception:
                pass

        return result

    except Exception as e:
        print(f"Warning: Could not read {filename}: {e}")
        return None


def load_all_data(data_dir):
    """
    Load ALL production Excel files from data_dir.
    Returns:
    {
        "months": ["2026-01", "2026-02"],   # sorted list of available months
        "data": {
            "2026-01": {
                "PL1": {...}, "PL2": {...}, ...
            },
            "2026-02": {
                "MIXER": {...}, ...
            }
        }
    }
    """
    all_entries = []

    for config_key, config in FILE_CONFIGS.items():
        # Search in data_dir and all subdirectories
        pattern_flat = os.path.join(data_dir, config["pattern"])
        pattern_sub = os.path.join(data_dir, "**", config["pattern"])
        files = sorted(set(
            glob.glob(pattern_flat) + glob.glob(pattern_sub, recursive=True)
        ))

        for filepath in files:
            # Skip temp files
            if os.path.basename(filepath).startswith('~$'):
                continue
            entry = _parse_file(filepath, config)
            if entry:
                all_entries.append(entry)

    # Group by month_key
    data_by_month = {}
    for entry in all_entries:
        mk = entry["month_key"]
        if mk not in data_by_month:
            data_by_month[mk] = {}
        data_by_month[mk][entry["name"]] = entry

    # Sort months descending (newest first)
    sorted_months = sorted(data_by_month.keys(), reverse=True)

    return {
        "months": sorted_months,
        "data": data_by_month,
    }


def get_latest_month_data(data_dir):
    """Load data and return only the latest month."""
    result = load_all_data(data_dir)
    if result["months"]:
        latest = result["months"][0]
        return result["data"][latest], latest
    return {}, None


def _find_pl_file(data_dir, pl_prefix, month, year):
    """Find the Excel file for a PL line (e.g., PL1 3.2026.xlsx).
    Searches in data_dir, then subdirectories recursively.
    """
    name_pattern = f"{pl_prefix} {month}.{year}.xls*"
    # Search in data_dir first
    files = glob.glob(os.path.join(data_dir, name_pattern))
    if files:
        return files[0]
    # Then search recursively in all subdirectories
    files = glob.glob(os.path.join(data_dir, "**", name_pattern), recursive=True)
    # Filter out temp files
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    return files[0] if files else None


def load_nvvh_for_day(data_dir, month, year, day):
    """Read NVVH (operator names) from PL1, PL6, and MIXER files for a specific day.

    PL1-5: AH49 (Ca1), AL49 (Ca2), AO49 (Ca3) from PL1 file, sheet named '{day}'.
    PL6-7: Same cells from PL6 file. 1 operator runs both PL6 and PL7.
    MIXER: M85 (Ca1), P85 (Ca2), S85 (Ca3) from MIXER file, sheet named '{day}'.
           Format: "CA1:name1 name2", "CA2:name1 + name2", "CA 3:name"
    Names are split by '+' or '-' delimiters.
    Returns: {
        "pl1_5": {"ca1": ["name1","name2","name3"], "ca2": [...], "ca3": [...]},
        "pl6_7": {"ca1": "name", "ca2": "name", "ca3": "name"},
        "mixer": {"ca1": "name", "ca2": "name", "ca3": "name"},
    }
    """
    import openpyxl

    result = {
        "pl1_5": {"ca1": [], "ca2": [], "ca3": []},
        "pl6_7": {"ca1": "", "ca2": "", "ca3": ""},
        "mixer": {"ca1": "", "ca2": "", "ca3": ""},
    }

    sheet_name = str(day)

    def _read_nvvh_row(filepath, row_num, col_indices):
        """Read a specific row from a specific sheet, return ca values."""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb[sheet_name]
            vals = {}
            for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False):
                for ca_key, col_idx in col_indices.items():
                    if col_idx < len(row):
                        vals[ca_key] = row[col_idx].value
            wb.close()
            return vals
        except Exception as e:
            print(f"Warning: Could not read NVVH from {filepath}: {e}")
            return {}

    # PL columns: AH=33, AL=37, AO=40 (0-based), row 49
    pl_ca_cols = {"ca1": 33, "ca2": 37, "ca3": 40}

    # PL1 file — for PL1-PL5 (3 operators for 5 machines)
    pl1_file = _find_pl_file(data_dir, "PL1", month, year)
    if pl1_file:
        vals = _read_nvvh_row(pl1_file, 49, pl_ca_cols)
        for ca_key in ["ca1", "ca2", "ca3"]:
            raw = vals.get(ca_key)
            if raw:
                names = re.split(r'[+\-]', str(raw))
                result["pl1_5"][ca_key] = [n.strip() for n in names if n.strip()]

    # PL6 file — for PL6-PL7 (1 operator for 2 machines)
    pl6_file = _find_pl_file(data_dir, "PL6", month, year)
    if pl6_file:
        vals = _read_nvvh_row(pl6_file, 49, pl_ca_cols)
        for ca_key in ["ca1", "ca2", "ca3"]:
            raw = vals.get(ca_key)
            result["pl6_7"][ca_key] = str(raw).strip() if raw else ""

    # MIXER file — row 85, columns M(12), P(15), S(18) (0-based)
    mixer_ca_cols = {"ca1": 12, "ca2": 15, "ca3": 18}
    mixer_file = _find_pl_file(data_dir, "MIXER T", month, year)
    if not mixer_file:
        # Try alternate naming: MIXER T{month}.{year}.xlsx
        import glob as _glob
        for pattern in [f"MIXER*T{month}.{year}.*", f"MIXER*{month}.{year}.*"]:
            files = _glob.glob(os.path.join(data_dir, "**", pattern), recursive=True)
            files = [f for f in files if not os.path.basename(f).startswith('~$')]
            if files:
                mixer_file = files[0]
                break
    if mixer_file:
        vals = _read_nvvh_row(mixer_file, 85, mixer_ca_cols)
        for ca_key in ["ca1", "ca2", "ca3"]:
            raw = vals.get(ca_key)
            if raw:
                name_str = str(raw).strip()
                # Strip "CA1:", "CA2:", "CA 3:" prefix
                name_str = re.sub(r'^CA\s*\d\s*[:：]\s*', '', name_str)
                # Split by + delimiter
                parts = re.split(r'[+]', name_str)
                result["mixer"][ca_key] = ','.join(n.strip() for n in parts if n.strip())

    return result


def _read_loss_sheet(filepath, day):
    """Read LOSS sheet from an Excel file for a specific day.
    Uses read_only mode for performance (files are ~2.5MB each).
    Returns list of loss entries with data > 0.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "LOSS" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["LOSS"]
        start_col = 5 + (day - 1) * 6  # 1-indexed column
        results = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=7, max_row=450,
                                                     values_only=False), start=7):
            # Column B (index 1) = code, Column C (index 2) = description
            code_cell = row[1] if len(row) > 1 else None
            desc_cell = row[2] if len(row) > 2 else None
            code = code_cell.value if code_cell else None
            if code is None:
                continue

            desc = desc_cell.value if desc_cell else ""

            # Read 6 data columns for this day
            vals = []
            for offset in range(6):
                col_idx = start_col - 1 + offset  # 0-indexed for row tuple
                v = row[col_idx].value if col_idx < len(row) else None
                vals.append(float(v) if v else 0)

            if any(v > 0 for v in vals):
                ca1_count, ca1_time, ca2_count, ca2_time, ca3_count, ca3_time = vals
                results.append({
                    "code": str(code),
                    "desc": str(desc) if desc else "",
                    "ca1_count": int(ca1_count),
                    "ca1_time": round(ca1_time * 60),
                    "ca2_count": int(ca2_count),
                    "ca2_time": round(ca2_time * 60),
                    "ca3_count": int(ca3_count),
                    "ca3_time": round(ca3_time * 60),
                })

        wb.close()
        return results
    except Exception as e:
        print(f"Warning: Could not read LOSS from {filepath}: {e}")
        return []


def load_loss_for_day(data_dir, month, year, day):
    """Read LOSS sheets from all PL files (PL1-PL7) and MIXER.

    Each PL file has its own LOSS sheet with per-machine data.
    MIXER LOSS uses pl_num=0.
    Returns: {"0": [...], "1": [...], "2": [...], ..., "7": [...]}
    """
    result = {}

    # PL1-PL7
    for pl_num in range(1, 8):
        pl_file = _find_pl_file(data_dir, f"PL{pl_num}", month, year)
        if pl_file:
            result[str(pl_num)] = _read_loss_sheet(pl_file, day)
        else:
            result[str(pl_num)] = []

    # MIXER (pl_num=0)
    mixer_file = _find_pl_file(data_dir, "MIXER T", month, year)
    if not mixer_file:
        import glob as _glob
        for pattern in [f"MIXER*T{month}.{year}.*", f"MIXER*{month}.{year}.*"]:
            files = _glob.glob(os.path.join(data_dir, "**", pattern), recursive=True)
            files = [f for f in files if not os.path.basename(f).startswith('~$')]
            if files:
                mixer_file = files[0]
                break
    if mixer_file:
        result["0"] = _read_loss_sheet(mixer_file, day)
    else:
        result["0"] = []

    return result


if __name__ == "__main__":
    result = load_all_data(os.path.dirname(os.path.abspath(__file__)))
    print(f"Available months: {result['months']}")
    for mk in result["months"]:
        lines = result["data"][mk]
        total = sum(info["summary"]["total"] for info in lines.values())
        m, y = mk.split("-")
        print(f"\nTháng {int(m)}/{y} — {len(lines)} dây chuyền — Tổng: {total:,.1f} tấn")
        for ln, info in lines.items():
            print(f"  {ln}: {info['summary']['total']:,.1f} tấn ({sum(1 for d in info['days'] if d['total'] > 0)} ngày)")
