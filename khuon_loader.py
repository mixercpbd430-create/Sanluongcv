"""
Khuôn (Mold) data loader — reads SERI sheet from PL1-PL7 Excel files.
Each PL file has a SERI sheet with:
  - Seri khuôn (mold serial number)
  - Thông số khuôn (mold specs)
  - Daily production (day 1-31)
  - TỔNG THÁNG / TỒN TRƯỚC / TỔNG summary columns

Column layout differs by PL:
  PL1, PL3-PL7: seri=B(1), khuon=C(2), day1=D(3)
  PL2:          seri=C(2), khuon=D(3), day1=E(4)
"""

import os
import re
import glob
import openpyxl


# Column offsets per PL (0-indexed)
PL_COLUMN_CONFIG = {
    1: {"seri": 1, "khuon": 2, "day1": 3},
    2: {"seri": 2, "khuon": 3, "day1": 4},
    3: {"seri": 1, "khuon": 2, "day1": 3},
    4: {"seri": 1, "khuon": 2, "day1": 3},
    5: {"seri": 1, "khuon": 2, "day1": 3},
    6: {"seri": 1, "khuon": 2, "day1": 3},
    7: {"seri": 1, "khuon": 2, "day1": 3},
}


def _find_pl_file(data_dir, pl_num, month, year):
    """Find the Excel file for a PL line, e.g., PL1 4.2026.xlsx."""
    name_pattern = f"PL{pl_num} {month}.{year}.xlsx"
    # Search in Update/PL{n} {year}/ first
    sub_dir = os.path.join(data_dir, "Update", f"PL{pl_num} {year}")
    filepath = os.path.join(sub_dir, name_pattern)
    if os.path.exists(filepath):
        return filepath
    # Fallback: search recursively
    files = glob.glob(os.path.join(data_dir, "**", name_pattern), recursive=True)
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    return files[0] if files else None


def load_khuon_data(data_dir, pl_num, month, year):
    """Read SERI sheet from one PL file.

    Returns list of molds:
    [
        {
            "seri": "23-3073 JM181",
            "thong_so": "2.5/65-20",
            "days": {1: 0, 2: 0, ..., 31: 0},
            "tong_thang": 515.2,
            "ton_truoc": 2501.6,
            "tong": 3016.8,
        },
        ...
    ]
    """
    filepath = _find_pl_file(data_dir, pl_num, month, year)
    if not filepath:
        return []

    config = PL_COLUMN_CONFIG.get(pl_num, {"seri": 1, "khuon": 2, "day1": 3})
    seri_idx = config["seri"]
    khuon_idx = config["khuon"]
    day1_idx = config["day1"]

    # Summary columns are always: day1 + 31, day1 + 32, day1 + 33
    tong_thang_idx = day1_idx + 31
    ton_truoc_idx = day1_idx + 32
    tong_idx = day1_idx + 33

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "SERI" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["SERI"]
        molds = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=3, max_row=100, min_col=1,
                         max_col=tong_idx + 1, values_only=True),
            start=3,
        ):
            # Check if seri column has a value
            seri_val = row[seri_idx] if len(row) > seri_idx else None
            if seri_val is None or str(seri_val).strip() == "":
                continue

            thong_so = row[khuon_idx] if len(row) > khuon_idx else ""

            # Read daily values (day 1-31)
            days = {}
            for d in range(31):
                col_idx = day1_idx + d
                val = row[col_idx] if len(row) > col_idx else 0
                days[d + 1] = round(float(val), 2) if val else 0

            # Read summary columns
            tong_thang = row[tong_thang_idx] if len(row) > tong_thang_idx else 0
            ton_truoc = row[ton_truoc_idx] if len(row) > ton_truoc_idx else 0
            tong = row[tong_idx] if len(row) > tong_idx else 0

            molds.append({
                "seri": str(seri_val).strip(),
                "thong_so": str(thong_so).strip() if thong_so else "",
                "days": days,
                "tong_thang": round(float(tong_thang), 2) if tong_thang else 0,
                "ton_truoc": round(float(ton_truoc), 2) if ton_truoc else 0,
                "tong": round(float(tong), 2) if tong else 0,
            })

        wb.close()
        return molds

    except Exception as e:
        print(f"Warning: Could not read SERI from PL{pl_num} {month}.{year}: {e}")
        return []


def load_all_khuon_for_month(data_dir, month, year):
    """Read SERI from all PL1-PL7 for a specific month.

    Returns: {
        "PL1": [mold1, mold2, ...],
        "PL2": [...],
        ...
    }
    """
    result = {}
    for pl_num in range(1, 8):
        molds = load_khuon_data(data_dir, pl_num, month, year)
        result[f"PL{pl_num}"] = molds
    return result


def load_khuon_yearly(data_dir, year):
    """Read SERI data for all months in a year, returning monthly totals per mold.

    Returns: {
        "PL1": [
            {
                "seri": "23-3073 JM181",
                "thong_so": "2.5/65-20",
                "months": {1: 0, 2: 120, 3: 515.2, ...},
                "year_total": 635.2,
            },
            ...
        ],
        ...
    }
    """
    result = {}

    for pl_num in range(1, 8):
        pl_key = f"PL{pl_num}"
        # Track molds by seri
        mold_map = {}

        for month in range(1, 13):
            molds = load_khuon_data(data_dir, pl_num, month, year)
            for m in molds:
                seri = m["seri"]
                if seri not in mold_map:
                    mold_map[seri] = {
                        "seri": seri,
                        "thong_so": m["thong_so"],
                        "months": {},
                        "year_total": 0,
                    }
                tong_thang = m["tong_thang"]
                if tong_thang > 0:
                    mold_map[seri]["months"][month] = tong_thang
                    mold_map[seri]["year_total"] += tong_thang

        # Convert to list, sorted by seri
        result[pl_key] = sorted(mold_map.values(), key=lambda x: x["seri"])

    return result


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    data_dir = os.path.dirname(os.path.abspath(__file__))

    # Test monthly load
    print("=== Monthly: PL1 April 2026 ===")
    molds = load_khuon_data(data_dir, 1, 4, 2026)
    for m in molds:
        active_days = sum(1 for v in m["days"].values() if v > 0)
        print(f"  {m['seri']} | {m['thong_so']} | "
              f"Tổng tháng: {m['tong_thang']}, Tồn trước: {m['ton_truoc']}, "
              f"Tổng: {m['tong']} | {active_days} ngày chạy")

    print(f"\n=== All PL - April 2026 ===")
    all_data = load_all_khuon_for_month(data_dir, 4, 2026)
    for pl, molds in all_data.items():
        total = sum(m["tong_thang"] for m in molds)
        print(f"  {pl}: {len(molds)} khuôn, tổng tháng = {total:,.1f} tấn")
